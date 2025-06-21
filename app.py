#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
南昌新东方凭证打印系统
Version: 2.4.0
Release Date: 2025-06-20

更新日志:
v2.4.0 (2025-06-20)
- 全面优化并发安全机制，支持多用户同时操作
- 新增并发处理工具模块和线程锁管理
- 优化数据库连接池配置，提升系统性能
- 添加死锁检测和自动重试机制

v2.2.0 (2025-06-19)
- 新增Excel批量导入用户功能，支持用户名、密码、姓名、邮箱、角色等完整信息导入
- 提供详细的Excel模板下载，包含数据模板和填写说明两个工作表
- 优化用户创建界面，Excel导入功能置于页面顶部突出推荐使用
- 增强数据验证机制，支持用户名唯一性检查、邮箱格式验证、角色权限验证
- 使用openpyxl库替代pandas，简化依赖并提升Excel处理性能
- 用户创建成功页面增加角色显示列，提供更完整的用户信息展示

v2.1.0 (2025-06-19)
- 新增消息盒子系统，支持打印成功、变更申请等各种消息通知
- 增强管理员审批功能，支持备注信息
- 优化用户信息变更流程，增加详细的状态反馈
- 改进操作员显示逻辑，智能判断显示用户姓名或用户名
- 完善用户信息管理，支持姓名和邮箱字段的搜索排序
- 增加消息类型字段长度至100字符，支持更长的消息类型名称

v2.0.0 (2025-06-18)
- 新增用户信息变更审批系统
- 添加姓名和邮箱字段支持
- 实现分级权限管理（管理员直接修改，普通用户需审批）
- 完善用户管理界面，支持搜索和排序功能

v1.0.0 (2025-06-01)
- 基础凭证打印功能
- 用户认证和权限管理
- 打印记录管理
- Cookies配置管理
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import json
import os
import secrets
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from utils import (
    ProofPrintSimulator, 
    TEMPLATE_MAPPING,
    CERTIFICATE_TYPES,
    print_certificate_by_biz_type,
    get_available_certificates
)
from utils.time_utils import get_beijing_time_str, get_beijing_timestamp, get_beijing_datetime
# 新的凭证管理器
from utils.certificate_manager import (
    generate_certificate_by_type,
    get_available_certificates as get_new_certificates
)
import base64
from io import BytesIO
from config import config
import pymysql
# 验证码相关导入
from PIL import Image, ImageDraw, ImageFont
import random
import string
import re  # 新增正则表达式模块用于密码验证

# 定时任务相关导入
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
import atexit
import threading
import uuid
from sqlalchemy import text
from contextlib import contextmanager

# 添加更多专用锁
_user_operation_lock = threading.Lock()
_message_operation_lock = threading.Lock()
_cookies_operation_lock = threading.Lock()

# 添加线程锁用于关键操作
_print_lock = threading.Lock()
_user_creation_lock = threading.Lock()
_cookies_config_lock = threading.Lock()

# 安装PyMySQL作为MySQLdb的替代
pymysql.install_as_MySQLdb()

app = Flask(__name__)

# 从环境变量获取配置模式，默认为development
config_name = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config.get(config_name, config['default']))

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录才能访问此页面'

# 初始化定时任务调度器
scheduler = BackgroundScheduler()
scheduler.start()
# 确保应用退出时停止调度器
atexit.register(lambda: scheduler.shutdown())

# 数据库模型
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    name = db.Column(db.String(100), nullable=True, comment='用户真实姓名')  # 新增姓名字段
    email = db.Column(db.String(120), nullable=True, comment='用户邮箱')  # 新增邮箱字段
    role = db.Column(db.String(20), nullable=False, default='user')  # 'admin' or 'user'
    is_enabled = db.Column(db.Boolean, default=True)
    is_first_login = db.Column(db.Boolean, default=True)  # 是否是首次登录
    created_at = db.Column(db.DateTime, default=get_beijing_datetime)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    # 软删除字段
    is_deleted = db.Column(db.Boolean, default=False, nullable=False)  # 软删除标记
    
    print_logs = db.relationship('PrintLog', backref='user', lazy=True)

class PrintLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    student_code = db.Column(db.String(50), nullable=False)
    student_name = db.Column(db.String(100), nullable=False)
    biz_type = db.Column(db.Integer, nullable=False)
    biz_name = db.Column(db.String(50), nullable=False)
    print_time = db.Column(db.DateTime, default=get_beijing_datetime)
    print_data = db.Column(db.Text, nullable=False)
    # 新增字段用于列表展示
    detail_info = db.Column(db.String(200), nullable=True)  # 详细信息，如"类型：充值，金额：¥1000.00"
    # 软删除字段
    is_deleted = db.Column(db.Boolean, default=False, nullable=False)  # 软删除标记

class CookiesConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, default='ERP Cookies')
    cookies_data = db.Column(db.Text, nullable=False)  # JSON格式存储cookies
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=get_beijing_datetime)
    updated_at = db.Column(db.DateTime, default=get_beijing_datetime, onupdate=get_beijing_datetime)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    last_test_time = db.Column(db.DateTime, nullable=True)  # 最后测试时间
    test_status = db.Column(db.String(20), default='未测试')  # 测试状态：成功、失败、未测试

class CookiesAutoCheck(db.Model):
    """Cookies自动检测配置模型"""
    id = db.Column(db.Integer, primary_key=True)
    is_enabled = db.Column(db.Boolean, default=False)  # 是否启用自动检测
    check_interval = db.Column(db.Integer, default=30)  # 检测间隔（分钟）
    failure_notification_sent = db.Column(db.Boolean, default=False)  # 是否已发送失败通知
    last_check_time = db.Column(db.DateTime, nullable=True)  # 最后检测时间
    consecutive_failures = db.Column(db.Integer, default=0)  # 连续失败次数
    created_at = db.Column(db.DateTime, default=get_beijing_datetime)
    updated_at = db.Column(db.DateTime, default=get_beijing_datetime, onupdate=get_beijing_datetime)

class UserChangeRequest(db.Model):
    """用户信息变更请求模型"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    request_type = db.Column(db.String(20), nullable=False)  # 'profile' 个人信息变更
    
    # 变更字段
    new_username = db.Column(db.String(80), nullable=True)  # 新用户名
    new_name = db.Column(db.String(100), nullable=True)     # 新姓名
    new_email = db.Column(db.String(120), nullable=True)    # 新邮箱
    
    # 原始字段（用于对比）
    old_username = db.Column(db.String(80), nullable=True)  # 原用户名
    old_name = db.Column(db.String(100), nullable=True)     # 原姓名
    old_email = db.Column(db.String(120), nullable=True)    # 原邮箱
    
    # 请求状态
    status = db.Column(db.String(20), default='pending', nullable=False)  # pending, approved, rejected
    reason = db.Column(db.Text, nullable=True)  # 申请理由
    admin_comment = db.Column(db.Text, nullable=True)  # 管理员备注
    
    # 时间字段
    created_at = db.Column(db.DateTime, default=get_beijing_datetime)
    processed_at = db.Column(db.DateTime, nullable=True)  # 处理时间
    processed_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # 处理人
    
    # 关联关系
    user = db.relationship('User', foreign_keys=[user_id], backref='change_requests')
    processor = db.relationship('User', foreign_keys=[processed_by])

class Message(db.Model):
    """用户消息模型"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # 消息类型
    message_type = db.Column(db.String(100), nullable=False)  # 'print_success', 'change_request_submitted', 'change_request_approved', 'change_request_rejected'
    
    # 消息内容
    title = db.Column(db.String(200), nullable=False)  # 消息标题
    content = db.Column(db.Text, nullable=False)  # 消息内容
    
    # 关联的对象ID（可选）
    related_id = db.Column(db.Integer, nullable=True)  # 关联的对象ID，如变更请求ID、打印记录ID等
    related_type = db.Column(db.String(50), nullable=True)  # 关联对象类型：'change_request', 'print_log'
    
    # 消息状态
    is_read = db.Column(db.Boolean, default=False, nullable=False)  # 是否已读
    
    # 时间字段
    created_at = db.Column(db.DateTime, default=get_beijing_datetime)
    read_at = db.Column(db.DateTime, nullable=True)  # 阅读时间
    
    # 关联关系
    user = db.relationship('User', backref='messages')

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# 密码安全验证函数
def validate_password_security(password):
    """
    验证密码安全性和防止危险字符
    返回: (is_valid, error_message)
    """
    if not password:
        return False, "密码不能为空"
    
    # 长度检查
    if len(password) < 6:
        return False, "密码长度至少6位"
    
    if len(password) > 128:
        return False, "密码长度不能超过128位"
    
    # 禁止的危险字符和SQL注入关键词
    dangerous_chars = [
        "'", '"', '\\', ';', '--', '/*', '*/', 'xp_', 'sp_',
        '<script', '</script>', '<', '>', '&lt;', '&gt;',
        'javascript:', 'vbscript:', 'onload=', 'onerror=',
        'eval(', 'exec(', 'system(', 'cmd(', '|', '&', '`'
    ]
    
    # SQL注入关键词检查（不区分大小写）
    sql_keywords = [
        'select', 'insert', 'update', 'delete', 'drop', 'create',
        'alter', 'exec', 'execute', 'union', 'or 1=1', 'and 1=1',
        'information_schema', 'sysobjects', 'syscolumns', 'master',
        'xp_cmdshell', 'sp_executesql', 'openrowset', 'opendatasource'
    ]
    
    password_lower = password.lower()
    
    # 检查危险字符
    for char in dangerous_chars:
        if char in password:
            return False, f"密码不能包含危险字符: {char}"
    
    # 检查SQL关键词
    for keyword in sql_keywords:
        if keyword in password_lower:
            return False, f"密码不能包含SQL关键词: {keyword}"
    
    # 检查是否包含HTML/JavaScript标签
    html_pattern = r'<[^>]*>'
    if re.search(html_pattern, password, re.IGNORECASE):
        return False, "密码不能包含HTML标签"
    
    # 检查是否全是相同字符
    if len(set(password)) == 1:
        return False, "密码不能全是相同字符"
    
    # 检查是否是连续字符（如123456, abcdef）
    consecutive_nums = '0123456789' * 2
    consecutive_chars = 'abcdefghijklmnopqrstuvwxyz' * 2
    if password.lower() in consecutive_nums or password.lower() in consecutive_chars:
        return False, "密码不能是连续的数字或字母"
    
    # 检查常见弱密码
    weak_passwords = [
        '123456', '123456789', 'password', 'admin', 'root', 'user',
        'guest', 'test', '111111', '000000', 'qwerty', 'abc123',
        'admin123', 'root123', 'user123', '123123', '321321',
        'password123', 'admin888', 'root888', '666666', '888888'
    ]
    
    if password.lower() in weak_passwords:
        return False, "密码过于简单，请使用更安全的密码"
    
    return True, "密码验证通过"

def sanitize_password_input(password):
    """
    清理密码输入，移除首尾空白字符
    """
    if not password:
        return ""
    
    # 只移除首尾空白，保留密码中间的空格（如果用户故意设置）
    return password.strip()

def check_password_strength(password):
    """
    检查密码强度
    返回: (strength_level, strength_text, suggestions)
    """
    if not password:
        return 0, "无", ["请输入密码"]
    
    score = 0
    suggestions = []
    
    # 长度检查
    if len(password) >= 8:
        score += 2
    elif len(password) >= 6:
        score += 1
    else:
        suggestions.append("密码长度至少6位，建议8位以上")
    
    # 包含小写字母
    if re.search(r'[a-z]', password):
        score += 1
    else:
        suggestions.append("建议包含小写字母")
    
    # 包含大写字母
    if re.search(r'[A-Z]', password):
        score += 1
    else:
        suggestions.append("建议包含大写字母")
    
    # 包含数字
    if re.search(r'[0-9]', password):
        score += 1
    else:
        suggestions.append("建议包含数字")
    
    # 包含特殊字符（安全的特殊字符）
    safe_special_chars = r'[!@#$%^&*()_+\-=\[\]{}|:,.<>?~]'
    if re.search(safe_special_chars, password):
        score += 2
    else:
        suggestions.append("建议包含特殊字符（如!@#$%等）")
    
    # 字符多样性
    unique_chars = len(set(password))
    if unique_chars >= len(password) * 0.7:
        score += 1
    
    # 确定强度等级
    if score >= 7:
        return 5, "很强", suggestions
    elif score >= 5:
        return 4, "强", suggestions
    elif score >= 3:
        return 3, "中等", suggestions
    elif score >= 1:
        return 2, "弱", suggestions
    else:
        return 1, "很弱", suggestions

def create_message(user_id, message_type, title, content, related_id=None, related_type=None):
    """创建用户消息"""
    try:
        message = Message(
            user_id=user_id,
            message_type=message_type,
            title=title,
            content=content,
            related_id=related_id,
            related_type=related_type
        )
        db.session.add(message)
        db.session.commit()
        return message  # 返回消息对象而不是True
    except Exception as e:
        print(f"创建消息失败: {e}")
        db.session.rollback()
        return None  # 返回None而不是False

def send_cookies_expired_notification():
    """向所有管理员发送cookies过期通知"""
    try:
        # 获取所有管理员用户
        admin_users = User.query.filter_by(role='admin', is_deleted=False, is_enabled=True).all()
        
        for admin in admin_users:
            # 检查是否已经发送过类似的通知（避免重复发送）
            existing_notification = Message.query.filter_by(
                user_id=admin.id,
                message_type='cookies_expired',
                is_read=False
            ).first()
            
            if not existing_notification:
                create_message(
                    user_id=admin.id,
                    message_type='cookies_expired',
                    title='系统认证配置已过期',
                    content=f'系统检测到当前的ERP认证配置（Cookies）已失效，请及时更新配置以确保系统正常运行。\n\n检测时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n\n请前往"Cookies配置"页面更新认证信息。',
                    related_type='cookies_config'
                )
        
        print(f"已向 {len(admin_users)} 位管理员发送cookies过期通知")
        return True
    except Exception as e:
        print(f"发送cookies过期通知失败: {e}")
        return False

def auto_check_cookies():
    """自动检测cookies状态"""
    try:
        with app.app_context():
            # 获取自动检测配置
            auto_check_config = CookiesAutoCheck.query.first()
            if not auto_check_config or not auto_check_config.is_enabled:
                return
            
            # 获取当前活跃的cookies配置
            active_config = CookiesConfig.query.filter_by(is_active=True).first()
            if not active_config:
                print("没有找到活跃的cookies配置")
                return
            
            print(f"开始自动检测cookies状态 - {get_beijing_datetime().strftime('%Y-%m-%d %H:%M:%S')}")
            
            try:
                cookies_dict = json.loads(active_config.cookies_data)
                
                # 使用测试学员编号测试API
                from utils.certificate_processors.search_student_certificate import search_student
                test_result = search_student(cookies_dict, None, 'NC24048S6UzC')
                
                # 更新检测时间
                auto_check_config.last_check_time = get_beijing_datetime()
                active_config.last_test_time = get_beijing_datetime()
                
                if test_result == 404:
                    # API请求失败，可能是cookies过期
                    active_config.test_status = '失败'
                    auto_check_config.consecutive_failures += 1
                    
                    print(f"Cookies检测失败，连续失败次数: {auto_check_config.consecutive_failures}")
                    
                    # 如果连续失败且未发送通知，则发送通知
                    if auto_check_config.consecutive_failures >= 2 and not auto_check_config.failure_notification_sent:
                        send_cookies_expired_notification()
                        auto_check_config.failure_notification_sent = True
                        print("已发送cookies过期通知")
                    
                elif test_result == 0 or isinstance(test_result, dict):
                    # 测试成功
                    active_config.test_status = '成功' 
                    auto_check_config.consecutive_failures = 0
                    auto_check_config.failure_notification_sent = False
                    print("Cookies检测成功")
                else:
                    # 其他情况视为失败
                    active_config.test_status = '失败'
                    auto_check_config.consecutive_failures += 1
                    print(f"Cookies检测失败（异常结果），连续失败次数: {auto_check_config.consecutive_failures}")
                
                db.session.commit()
                
            except Exception as e:
                print(f"执行cookies检测时出错: {e}")
                auto_check_config.consecutive_failures += 1
                db.session.commit()
    
    except Exception as e:
        print(f"自动检测cookies时出错: {e}")

def init_cookies_auto_check():
    """初始化cookies自动检测"""
    try:
        with app.app_context():
            # 检查是否已存在配置
            auto_check_config = CookiesAutoCheck.query.first()
            if not auto_check_config:
                # 创建默认配置
                auto_check_config = CookiesAutoCheck(
                    is_enabled=False,
                    check_interval=30
                )
                db.session.add(auto_check_config)
                db.session.commit()
                print("已创建默认的cookies自动检测配置")
            
            # 如果启用了自动检测，启动定时任务
            if auto_check_config.is_enabled:
                start_cookies_auto_check()
                
    except Exception as e:
        print(f"初始化cookies自动检测失败: {e}")

def start_cookies_auto_check():
    """启动cookies自动检测任务"""
    try:
        with app.app_context():
            auto_check_config = CookiesAutoCheck.query.first()
            if not auto_check_config or not auto_check_config.is_enabled:
                return
            
            # 移除现有的任务（如果存在）
            try:
                scheduler.remove_job('cookies_auto_check')
            except:
                pass
            
            # 添加新的定时任务
            scheduler.add_job(
                func=auto_check_cookies,
                trigger=IntervalTrigger(minutes=auto_check_config.check_interval),
                id='cookies_auto_check',
                name='Cookies自动检测任务',
                replace_existing=True
            )
            
            print(f"已启动cookies自动检测任务，检测间隔: {auto_check_config.check_interval} 分钟")
            
    except Exception as e:
        print(f"启动cookies自动检测任务失败: {e}")

def stop_cookies_auto_check():
    """停止cookies自动检测任务"""
    try:
        scheduler.remove_job('cookies_auto_check')
        print("已停止cookies自动检测任务")
    except:
        pass

# Cookies超时检查
@app.before_request
def check_cookie_timeout():
    """检查cookies是否超时，超时则强制退出"""
    # 排除不需要登录的路由
    excluded_endpoints = ['login', 'captcha', 'static']
    
    if request.endpoint in excluded_endpoints:
        return
    
    # 如果用户已登录，检查cookies超时和用户身份
    if current_user.is_authenticated:
        login_time_cookie = request.cookies.get('login_time')
        login_user_cookie = request.cookies.get('login_user')
        login_user_id_cookie = request.cookies.get('login_user_id')
        
        # 检查用户身份是否匹配
        if login_user_cookie and login_user_id_cookie:
            try:
                cookie_user_id = int(login_user_id_cookie)
                if (login_user_cookie != current_user.username or 
                    cookie_user_id != current_user.id):
                    # 用户身份不匹配，可能是会话劫持或用户切换
                    logout_user()
                    session.clear()
                    response = make_response(redirect(url_for('login')))
                    # 删除所有相关cookies
                    response.set_cookie('login_time', '', expires=0)
                    response.set_cookie('login_user', '', expires=0)
                    response.set_cookie('login_user_id', '', expires=0)
                    flash('登录身份验证失败，请重新登录', 'warning')
                    return response
            except (ValueError, TypeError):
                # cookie值无效，强制退出
                logout_user()
                session.clear()
                response = make_response(redirect(url_for('login')))
                response.set_cookie('login_time', '', expires=0)
                response.set_cookie('login_user', '', expires=0)
                response.set_cookie('login_user_id', '', expires=0)
                flash('登录状态异常，请重新登录', 'warning')
                return response
        
        # 检查登录时间
        if login_time_cookie:
            try:
                # 解析登录时间
                login_time = datetime.fromtimestamp(float(login_time_cookie))
                now = datetime.now()
                
                # 检查是否超过12小时
                if now - login_time > timedelta(hours=12):
                    # 超时，强制退出
                    logout_user()
                    session.clear()
                    response = make_response(redirect(url_for('login')))
                    # 删除所有相关cookies
                    response.set_cookie('login_time', '', expires=0)
                    response.set_cookie('login_user', '', expires=0)
                    response.set_cookie('login_user_id', '', expires=0)
                    flash('登录已超时（12小时），请重新登录', 'warning')
                    return response
            except (ValueError, TypeError):
                # cookie值无效，强制退出
                logout_user()
                session.clear()
                response = make_response(redirect(url_for('login')))
                response.set_cookie('login_time', '', expires=0)
                response.set_cookie('login_user', '', expires=0)
                response.set_cookie('login_user_id', '', expires=0)
                flash('登录状态异常，请重新登录', 'warning')
                return response
        else:
            # 没有登录时间cookie但用户显示为已认证，可能是状态不一致，强制退出重新登录
            logout_user()
            session.clear()
            response = make_response(redirect(url_for('login')))
            response.set_cookie('login_time', '', expires=0)
            response.set_cookie('login_user', '', expires=0)
            response.set_cookie('login_user_id', '', expires=0)
            flash('登录状态异常，请重新登录', 'warning')
            return response

# 验证码生成函数
def generate_captcha_code():
    """生成4位随机验证码"""
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))

def create_captcha_image(code):
    """创建验证码图片"""
    # 图片尺寸 - 增大尺寸以提高可读性
    width, height = 180, 80
    
    # 创建图片和绘图对象
    image = Image.new('RGB', (width, height), color='white')
    draw = ImageDraw.Draw(image)
    
    # 尝试使用系统字体，如果失败则使用默认字体
    try:
        # Linux 系统常见字体
        font = ImageFont.truetype('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', 36)
    except:
        try:
            # 备选字体
            font = ImageFont.truetype('/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf', 36)
        except:
            try:
                # Windows 系统常见字体
                font = ImageFont.truetype('arial.ttf', 36)
            except:
                try:
                    # 备选字体
                    font = ImageFont.truetype('calibri.ttf', 36)
                except:
                    # 使用默认字体并尝试获得更大的字体
                    try:
                        font = ImageFont.load_default()
                        # 如果支持的话，尝试加载更大的默认字体
                        font = ImageFont.truetype("arial.ttf", 36)
                    except:
                        font = ImageFont.load_default()
    
    # 绘制验证码文字 - 调整间距和位置
    for i, char in enumerate(code):
        x = 25 + i * 35  # 增加字符间距
        y = random.randint(15, 25)  # 调整垂直位置
        # 使用更清晰的颜色，避免太浅的颜色
        color = (random.randint(0, 80), random.randint(0, 80), random.randint(0, 80))
        # 可以添加轻微的旋转角度让验证码更有变化
        draw.text((x, y), char, font=font, fill=color)
    
    # 减少干扰线的数量，避免影响可读性
    for _ in range(3):
        x1 = random.randint(0, width)
        y1 = random.randint(0, height)
        x2 = random.randint(0, width)
        y2 = random.randint(0, height)
        # 使用更浅的颜色做干扰线
        draw.line([(x1, y1), (x2, y2)], fill=(random.randint(150, 200), random.randint(150, 200), random.randint(150, 200)), width=1)
    
    # 减少干扰点的数量
    for _ in range(30):
        x = random.randint(0, width)
        y = random.randint(0, height)
        # 使用更浅的颜色做干扰点
        draw.point((x, y), fill=(random.randint(180, 255), random.randint(180, 255), random.randint(180, 255)))
    
    return image

# 权限装饰器
def admin_required(f):
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('需要管理员权限才能访问此页面', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# 首次登录检查装饰器
def first_login_required(f):
    def decorated_function(*args, **kwargs):
        if current_user.is_authenticated and current_user.is_first_login:
            flash('请先修改您的密码后再使用系统功能', 'warning')
            return redirect(url_for('first_login_change_password'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# 路由
@app.route('/')
def index():
    # 检查登录状态和cookie一致性
    if current_user.is_authenticated:
        # 检查cookie是否存在且有效
        login_time_cookie = request.cookies.get('login_time')
        login_user_cookie = request.cookies.get('login_user')
        login_user_id_cookie = request.cookies.get('login_user_id')
        
        # 如果用户已登录但cookie不存在或不匹配，强制重新登录
        if not login_time_cookie or not login_user_cookie or not login_user_id_cookie:
            logout_user()
            session.clear()
            flash('登录状态异常，请重新登录', 'warning')
            return redirect(url_for('login'))
        
        # 检查cookie中的用户信息是否匹配
        try:
            cookie_user_id = int(login_user_id_cookie)
            if login_user_cookie != current_user.username or cookie_user_id != current_user.id:
                logout_user()
                session.clear()
                flash('登录身份验证失败，请重新登录', 'warning')
                return redirect(url_for('login'))
        except (ValueError, TypeError):
            logout_user()
            session.clear()
            flash('登录状态异常，请重新登录', 'warning')
            return redirect(url_for('login'))
        
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/captcha')
def captcha():
    """生成验证码图片"""
    # 生成验证码
    code = generate_captcha_code()
    
    # 将验证码存储到session中
    session['captcha_code'] = code.upper()
    
    # 创建验证码图片
    image = create_captcha_image(code)
    
    # 将图片转换为字节流
    img_io = BytesIO()
    image.save(img_io, 'PNG')
    img_io.seek(0)
    
    # 返回图片响应
    response = make_response(img_io.getvalue())
    response.headers['Content-Type'] = 'image/png'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    return response

@app.route('/login', methods=['GET', 'POST'])
def login():
    # 如果用户已经登录，重定向到dashboard
    if current_user.is_authenticated:
        # 检查cookie是否存在且有效
        login_time_cookie = request.cookies.get('login_time')
        login_user_cookie = request.cookies.get('login_user')
        login_user_id_cookie = request.cookies.get('login_user_id')
        
        # 如果cookie都存在且匹配，则重定向到dashboard
        if (login_time_cookie and login_user_cookie and login_user_id_cookie):
            try:
                cookie_user_id = int(login_user_id_cookie)
                if login_user_cookie == current_user.username and cookie_user_id == current_user.id:
                    return redirect(url_for('dashboard'))
            except (ValueError, TypeError):
                pass
        
        # 如果cookie不匹配或不存在，强制退出重新登录
        logout_user()
        session.clear()
        flash('登录状态异常，请重新登录', 'warning')
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        captcha_input = request.form.get('captcha', '').upper()
        
        # 验证验证码
        if 'captcha_code' not in session or captcha_input != session.get('captcha_code'):
            flash('验证码错误', 'error')
            # 清除旧的验证码
            session.pop('captcha_code', None)
            return render_template('login.html')
        
        # 清除验证码（一次性使用）
        session.pop('captcha_code', None)
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            if not user.is_enabled:
                flash('账号已被禁用，请联系管理员', 'error')
                return render_template('login.html')
            
            login_user(user)
            
            # 检查是否是首次登录
            if user.is_first_login:
                flash('这是您的首次登录，请立即修改您的密码以确保账户安全！', 'warning')
                response = make_response(redirect(url_for('first_login_change_password')))
            else:
                flash(f'欢迎回来，{user.username}！', 'success')
                response = make_response(redirect(url_for('dashboard')))
            
            # 设置登录时间cookie和用户绑定cookie，有效期12小时
            current_timestamp = str(datetime.now().timestamp())
            
            response.set_cookie('login_time', 
                              current_timestamp, 
                              max_age=12*60*60,  # 12小时（秒）
                              httponly=True,     # 防止JavaScript访问
                              secure=False)      # 开发环境设为False，生产环境应设为True
            
            response.set_cookie('login_user', 
                              user.username,     # 存储用户名
                              max_age=12*60*60,  # 12小时（秒）
                              httponly=True,     # 防止JavaScript访问
                              secure=False)      # 开发环境设为False，生产环境应设为True
            
            response.set_cookie('login_user_id', 
                              str(user.id),      # 存储用户ID作为双重验证
                              max_age=12*60*60,  # 12小时（秒）
                              httponly=True,     # 防止JavaScript访问
                              secure=False)      # 开发环境设为False，生产环境应设为True
            
            return response
        else:
            flash('用户名或密码错误', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    # 记录用户退出信息
    username = current_user.username if current_user.is_authenticated else '未知用户'
    
    # 强制退出用户
    logout_user()
    
    # 清除所有session数据，包括Flask-Login的session
    session.clear()
    
    # 如果有其他需要清除的session key，也在这里清除
    for key in list(session.keys()):
        session.pop(key, None)
    
    flash('已成功退出登录', 'info')
    
    # 创建响应并删除所有登录相关cookies
    response = make_response(redirect(url_for('login')))
    
    # 删除所有可能的登录相关cookies
    response.set_cookie('login_time', '', expires=0, path='/')
    response.set_cookie('login_user', '', expires=0, path='/')
    response.set_cookie('login_user_id', '', expires=0, path='/')
    
    # 也删除可能的Flask-Login session cookie
    response.set_cookie('session', '', expires=0, path='/')
    
    # 设置缓存控制头，确保页面不被缓存
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    return response

@app.route('/dashboard')
@login_required
@first_login_required
def dashboard():
    # 获取用户的打印记录统计
    user_print_count = PrintLog.query.filter_by(user_id=current_user.id).count()
    
    # 如果是管理员，获取更多统计信息
    stats = {
        'total_users': User.query.filter_by(is_deleted=False).count(),
        'active_users': User.query.filter_by(is_deleted=False, is_enabled=True).count(),
        'total_prints': PrintLog.query.count(),
        'recent_prints': PrintLog.query.order_by(PrintLog.print_time.desc()).limit(5).all()
    }
    
    return render_template('dashboard.html', user_print_count=user_print_count, stats=stats)

@app.route('/users')
@login_required
@first_login_required
@admin_required
def users():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    # 限制每页记录数量范围
    if per_page not in [10, 20, 50, 100]:
        per_page = 20
    
    # 搜索参数
    search_username = request.args.get('search_username', '').strip()
    search_name = request.args.get('search_name', '').strip()  # 新增姓名搜索
    search_email = request.args.get('search_email', '').strip()  # 新增邮箱搜索
    search_role = request.args.get('search_role', '').strip()
    search_status = request.args.get('search_status', '').strip()
    search_activation = request.args.get('search_activation', '').strip()
    search_date_start = request.args.get('search_date_start', '').strip()
    search_date_end = request.args.get('search_date_end', '').strip()
    search_creator = request.args.get('search_creator', '').strip()
    
    # 排序参数
    sort_by = request.args.get('sort_by', 'created_at')  # 默认按创建时间排序
    sort_order = request.args.get('sort_order', 'desc')  # 默认降序
    
    # 构建查询 - 只显示未删除的用户
    query = User.query.filter_by(is_deleted=False)
    
    # 用户名筛选
    if search_username:
        query = query.filter(User.username.like(f'%{search_username}%'))
    
    # 姓名筛选
    if search_name:
        query = query.filter(User.name.like(f'%{search_name}%'))
    
    # 邮箱筛选
    if search_email:
        query = query.filter(User.email.like(f'%{search_email}%'))
    
    # 角色筛选
    if search_role:
        query = query.filter(User.role == search_role)
    
    # 状态筛选（启用/禁用）
    if search_status:
        is_enabled = search_status == 'enabled'
        query = query.filter(User.is_enabled == is_enabled)
    
    # 激活状态筛选
    if search_activation:
        is_first_login = search_activation == 'pending'
        query = query.filter(User.is_first_login == is_first_login)
    
    # 时间范围筛选
    if search_date_start:
        try:
            start_date = datetime.strptime(search_date_start, '%Y-%m-%d')
            query = query.filter(User.created_at >= start_date)
        except ValueError:
            pass
    
    if search_date_end:
        try:
            end_date = datetime.strptime(search_date_end, '%Y-%m-%d')
            # 结束日期包含当天，所以加1天
            end_date = end_date + timedelta(days=1)
            query = query.filter(User.created_at < end_date)
        except ValueError:
            pass
    
    # 创建者筛选
    if search_creator:
        # 先查找创建者用户
        creator_query = User.query.filter(User.username.like(f'%{search_creator}%')).with_entities(User.id)
        creator_ids = [c.id for c in creator_query.all()]
        if creator_ids:
            query = query.filter(User.created_by.in_(creator_ids))
        else:
            # 如果没有找到匹配的创建者，返回空结果
            query = query.filter(User.id == -1)
    
    # 排序和分页
    # 构建排序
    valid_sort_fields = {
        'username': User.username,
        'name': User.name,
        'email': User.email,
        'role': User.role,
        'is_enabled': User.is_enabled,
        'is_first_login': User.is_first_login,
        'created_at': User.created_at,
        'id': User.id
    }
    
    sort_field = valid_sort_fields.get(sort_by, User.created_at)
    
    if sort_order == 'asc':
        query = query.order_by(sort_field.asc())
    else:
        query = query.order_by(sort_field.desc())
    
    users = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # 获取所有用户用于创建者显示（只获取未删除的）
    all_users = User.query.filter_by(is_deleted=False).all()
    
    # 构建搜索参数字典，用于分页链接保持搜索状态
    search_params = {}
    if search_username:
        search_params['search_username'] = search_username
    if search_name:
        search_params['search_name'] = search_name
    if search_email:
        search_params['search_email'] = search_email
    if search_role:
        search_params['search_role'] = search_role
    if search_status:
        search_params['search_status'] = search_status
    if search_activation:
        search_params['search_activation'] = search_activation
    if search_date_start:
        search_params['search_date_start'] = search_date_start
    if search_date_end:
        search_params['search_date_end'] = search_date_end
    if search_creator:
        search_params['search_creator'] = search_creator
    
    # 添加排序和分页参数
    search_params['sort_by'] = sort_by
    search_params['sort_order'] = sort_order
    search_params['per_page'] = per_page
    
    # 构建分页URL参数字符串
    search_query_string = ''
    if search_params:
        query_parts = []
        for key, value in search_params.items():
            if value:  # 只包含非空值
                query_parts.append(f'{key}={value}')
        if query_parts:
            search_query_string = '&' + '&'.join(query_parts)
    
    # 获取统计信息
    stats = {
        'total_users': User.query.filter_by(is_deleted=False).count(),
        'enabled_users': User.query.filter_by(is_deleted=False, is_enabled=True).count(),
        'disabled_users': User.query.filter_by(is_deleted=False, is_enabled=False).count(),
        'activated_users': User.query.filter_by(is_deleted=False, is_first_login=False).count(),
        'pending_users': User.query.filter_by(is_deleted=False, is_first_login=True).count(),
        'admin_users': User.query.filter_by(is_deleted=False, role='admin').count(),
        'regular_users': User.query.filter_by(is_deleted=False, role='user').count(),
    }
    
    return render_template('users.html', 
                         users=users, 
                         all_users=all_users,
                         search_params=search_params,
                         search_query_string=search_query_string,
                         current_per_page=per_page,
                         stats=stats)

@app.route('/change_requests')
@login_required
@first_login_required
@admin_required
def change_requests():
    """管理员查看变更请求"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status_filter = request.args.get('status', 'all').strip()
    
    # 构建查询
    query = UserChangeRequest.query
    
    # 状态筛选
    if status_filter and status_filter != 'all':
        query = query.filter(UserChangeRequest.status == status_filter)
    
    # 按创建时间倒序
    query = query.order_by(UserChangeRequest.created_at.desc())
    
    # 分页
    requests_page = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # 获取所有用户信息用于显示
    all_users = User.query.filter_by(is_deleted=False).all()
    user_dict = {user.id: user for user in all_users}
    
    # 统计信息
    stats = {
        'total': UserChangeRequest.query.count(),
        'pending': UserChangeRequest.query.filter_by(status='pending').count(),
        'approved': UserChangeRequest.query.filter_by(status='approved').count(),
        'rejected': UserChangeRequest.query.filter_by(status='rejected').count(),
    }
    
    return render_template('change_requests.html', 
                         requests=requests_page,
                         user_dict=user_dict,
                         status_filter=status_filter,
                         current_per_page=per_page,
                         stats=stats)

@app.route('/process_change_request/<int:request_id>/<action>', methods=['GET', 'POST'])
@login_required
@first_login_required
@admin_required
def process_change_request(request_id, action):
    """处理变更请求"""
    if action not in ['approve', 'reject']:
        flash('无效的操作', 'error')
        return redirect(url_for('change_requests'))
    
    # 使用线程锁确保变更请求处理的并发安全
    with _user_operation_lock:
        try:
            with safe_db_transaction() as session:
                change_request = UserChangeRequest.query.with_for_update().get_or_404(request_id)
                
                if change_request.status != 'pending':
                    flash('该请求已被处理', 'warning')
                    return redirect(url_for('change_requests'))
                
                # 从POST或GET获取管理员备注
                admin_comment = ''
                if request.method == 'POST':
                    admin_comment = request.form.get('admin_comment', '').strip()
                else:
                    admin_comment = request.args.get('comment', '').strip()
                
                if action == 'approve':
                    # 批准变更
                    user = User.query.with_for_update().get(change_request.user_id)
                    if not user:
                        flash('用户不存在', 'error')
                        return redirect(url_for('change_requests'))
                    
                    # 应用变更
                    if change_request.new_username:
                        # 再次检查用户名是否已存在
                        existing_user = User.query.filter_by(username=change_request.new_username, is_deleted=False).with_for_update().first()
                        if existing_user and existing_user.id != user.id:
                            flash('用户名已被其他用户使用，无法批准', 'error')
                            return redirect(url_for('change_requests'))
                        user.username = change_request.new_username
                    
                    if change_request.new_name is not None:
                        user.name = change_request.new_name
                    
                    if change_request.new_email is not None:
                        user.email = change_request.new_email
                    
                    # 更新请求状态
                    change_request.status = 'approved'
                    change_request.processed_at = get_beijing_datetime()
                    change_request.processed_by = current_user.id
                    change_request.admin_comment = admin_comment
                    
                    # 创建成功消息
                    changes = []
                    if change_request.new_username and change_request.new_username != change_request.old_username:
                        changes.append(f"用户名: {change_request.old_username} → {change_request.new_username}")
                    if change_request.new_name != change_request.old_name:
                        old_name = change_request.old_name or "未设置"
                        new_name = change_request.new_name or "未设置"
                        changes.append(f"姓名: {old_name} → {new_name}")
                    if change_request.new_email != change_request.old_email:
                        old_email = change_request.old_email or "未设置"
                        new_email = change_request.new_email or "未设置"
                        changes.append(f"邮箱: {old_email} → {new_email}")
                    
                    message_content = f"您的信息变更申请已被批准并生效。\n\n变更内容：\n" + "\n".join(changes)
                    if admin_comment:
                        message_content += f"\n\n管理员备注：{admin_comment}"
                    
                    message = Message(
                        user_id=change_request.user_id,
                        message_type='change_request_approved',
                        title='信息变更申请已批准',
                        content=message_content,
                        related_id=change_request.id,
                        related_type='change_request'
                    )
                    session.add(message)
                    
                    flash(f'已批准用户 {user.username} 的信息变更申请', 'success')
                    
                else:  # reject
                    change_request.status = 'rejected'
                    change_request.processed_at = get_beijing_datetime()
                    change_request.processed_by = current_user.id
                    change_request.admin_comment = admin_comment
                    
                    # 创建拒绝消息
                    changes = []
                    if change_request.new_username and change_request.new_username != change_request.old_username:
                        changes.append(f"用户名: {change_request.old_username} → {change_request.new_username}")
                    if change_request.new_name != change_request.old_name:
                        old_name = change_request.old_name or "未设置"
                        new_name = change_request.new_name or "未设置"
                        changes.append(f"姓名: {old_name} → {new_name}")
                    if change_request.new_email != change_request.old_email:
                        old_email = change_request.old_email or "未设置"
                        new_email = change_request.new_email or "未设置"
                        changes.append(f"邮箱: {old_email} → {new_email}")
                    
                    message_content = f"您的信息变更申请已被拒绝。\n\n申请内容：\n" + "\n".join(changes)
                    if admin_comment:
                        message_content += f"\n\n拒绝原因：{admin_comment}"
                    
                    message = Message(
                        user_id=change_request.user_id,
                        message_type='change_request_rejected',
                        title='信息变更申请已拒绝',
                        content=message_content,
                        related_id=change_request.id,
                        related_type='change_request'
                    )
                    session.add(message)
                    
                    flash('已拒绝该变更申请', 'info')
                
        except Exception as e:
            flash(f'处理失败：{str(e)}', 'error')
    
    return redirect(url_for('change_requests'))

@app.route('/create_user', methods=['GET', 'POST'])
@login_required
@admin_required
def create_user():
    if request.method == 'POST':
        # 使用线程锁确保用户创建操作的并发安全
        with _user_creation_lock:
            # 获取表单数据
            usernames = request.form.get('usernames', '').strip()
            passwords = request.form.get('passwords', '').strip()
            names = request.form.get('names', '').strip()
            emails = request.form.get('emails', '').strip()
            role = request.form.get('role', 'user')
            
            if not usernames:
                flash('请输入用户名', 'error')
                return render_template('create_user.html')
            
            # 分割输入的数据
            username_list = [u.strip() for u in usernames.split('\n') if u.strip()]
            password_list = [p.strip() for p in passwords.split('\n') if p.strip()] if passwords else []
            name_list = [n.strip() for n in names.split('\n') if n.strip()] if names else []
            email_list = [e.strip() for e in emails.split('\n') if e.strip()] if emails else []
            
            created_users = []
            errors = []
            
            # 使用安全的数据库事务上下文
            with safe_db_transaction() as session:
                for i, username in enumerate(username_list):
                    # 并发安全的用户名唯一性检查
                    if not check_username_uniqueness_safe(username):
                        errors.append(f'用户名 {username} 已存在')
                        continue
                    
                    # 如果提供了密码列表，使用对应的密码，否则使用默认密码
                    if i < len(password_list):
                        original_password = sanitize_password_input(password_list[i])
                        password = original_password
                        is_using_default_password = False
                    else:
                        password = '123456'  # 默认密码
                        is_using_default_password = True
                    
                    # 只有当管理员提供了密码时才进行密码安全性检查
                    # 如果使用默认密码，则跳过密码检查
                    if not is_using_default_password:
                        # 管理员提供了密码，需要进行安全性检查
                        is_valid, error_msg = validate_password_security(password)
                        if not is_valid:
                            errors.append(f'用户 {username} 的密码不符合安全要求：{error_msg}')
                            continue
                    
                    # 获取对应的姓名和邮箱
                    name = name_list[i] if i < len(name_list) else None
                    email = email_list[i] if i < len(email_list) else None
                    
                    try:
                        user = User(
                            username=username,
                            password_hash=generate_password_hash(password),
                            name=name,
                            email=email,
                            role=role,
                            created_by=current_user.id
                        )
                        session.add(user)
                        created_users.append({
                            'username': username, 
                            'password': password, 
                            'name': name, 
                            'email': email,
                            'role': role
                        })
                    except Exception as e:
                        errors.append(f'创建用户 {username} 失败：{str(e)}')
                        continue
                
                # 显示结果
                if created_users:
                    flash(f'成功创建 {len(created_users)} 个用户', 'success')
                    # 创建用户创建成功消息
                    user_list = '\n'.join([f"• {user['username']} ({user['name'] or '未设置姓名'})" for user in created_users])
                    create_message(
                        user_id=current_user.id,
                        message_type='users_created',
                        title=f'成功创建{len(created_users)}个用户',
                        content=f'您已成功批量创建{len(created_users)}个用户账号\n\n创建时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n\n用户列表：\n{user_list}\n\n请及时将账号信息安全地分发给相应用户。',
                        related_type='user_management'
                    )
                if errors:
                    for error in errors:
                        flash(error, 'warning')
                
                if created_users:
                    return render_template('user_created.html', users=created_users)
    
    return render_template('create_user.html')

@app.route('/download_user_template')
@login_required
@admin_required
def download_user_template():
    """下载用户批量导入Excel模板"""
    try:
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "用户批量导入模板"
        
        # 设置表头
        headers = ['用户名*', '密码', '姓名', '邮箱', '角色', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # 设置表头样式
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 添加示例数据
        sample_data = [
            ['user001', '123456', '张三', 'zhangsan@example.com', 'user', '示例用户1'],
            ['user002', '123456', '李四', 'lisi@example.com', 'user', '示例用户2'],
            ['user003', '123456', '王五', 'wangwu@example.com', 'admin', '示例管理员'],
            ['teacher001', '123456', '张老师', 'teacher@example.com', 'user', '教师账号'],
            ['student001', '123456', '小明', 'student@example.com', 'user', '学生账号']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # 为角色列添加颜色区分
                if col == 5:  # 角色列
                    if value == 'admin':
                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        
        # 设置列宽
        column_widths = [15, 12, 12, 25, 8, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 添加说明工作表
        ws_info = wb.create_sheet("填写说明")
        
        # 说明内容
        instructions = [
            ['字段名称', '是否必填', '说明', '示例'],
            ['用户名', '是', '用户登录名，必须唯一，不能重复', 'user001, teacher001'],
            ['密码', '否', '用户密码，如果不填写则使用默认密码：123456（无需验证）', 'Abc123456, MyPass@2024'],
            ['姓名', '否', '用户真实姓名，支持中文', '张三, 李老师'],
            ['邮箱', '否', '用户邮箱地址', 'user@example.com'],
            ['角色', '否', '用户角色，只能填写 user 或 admin，默认为 user', 'user, admin'],
            ['备注', '否', '备注信息，仅用于说明，不会导入到系统中', '任意文本'],
            ['', '', '', ''],
            ['重要说明：', '', '', ''],
            ['1. 带*号的字段为必填项', '', '', ''],
            ['2. 用户名不能重复，如果系统中已存在相同用户名，该行将被跳过', '', '', ''],
            ['3. 密码如果不填写，系统将使用默认密码：123456（无需安全性验证）', '', '', ''],
            ['4. 角色只能填写 user（普通用户）或 admin（管理员）', '', '', ''],
            ['5. 管理员角色拥有完整的系统管理权限，请谨慎分配', '', '', ''],
            ['6. 示例数据可以删除，填写您的实际数据', '', '', ''],
            ['7. 支持批量导入，一次最多导入1000个用户', '', '', ''],
            ['8. 请保持Excel格式不变，不要修改表头', '', '', ''],
            ['', '', '', ''],
            ['密码安全要求（仅对自定义密码）：', '', '', ''],
            ['• 长度至少6位，建议8位以上', '', '', ''],
            ['• 不能包含危险字符：引号、分号、脚本标签等', '', '', ''],
            ['• 不能包含SQL关键词：select、insert、delete等', '', '', ''],
            ['• 不能是常见弱密码：123456、password、admin等', '', '', ''],
            ['• 不能全是相同字符或连续字符', '', '', ''],
            ['• 建议包含大小写字母、数字和特殊字符', '', '', ''],
            ['• 密码为空时自动使用默认密码123456，无需验证', '', '', '']
        ]
        
        for row, data in enumerate(instructions, 1):
            for col, value in enumerate(data, 1):
                cell = ws_info.cell(row=row, column=col, value=value)
                if row == 1:  # 表头
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif row >= 9 and col == 1:  # 重要说明
                    cell.font = Font(bold=True, color='D32F2F')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # 设置说明工作表列宽
        info_column_widths = [15, 10, 50, 25]
        for col, width in enumerate(info_column_widths, 1):
            ws_info.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f'用户批量导入模板_{get_beijing_timestamp()}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'生成模板失败：{str(e)}', 'error')
        return redirect(url_for('create_user'))

@app.route('/import_users_excel', methods=['POST'])
@login_required
@admin_required
def import_users_excel():
    """Excel批量导入用户"""
    if 'excel_file' not in request.files:
        flash('请选择要上传的Excel文件', 'error')
        return redirect(url_for('create_user'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash('请选择要上传的Excel文件', 'error')
        return redirect(url_for('create_user'))
    
    if not file or not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('请上传有效的Excel文件（.xlsx或.xls格式）', 'error')
        return redirect(url_for('create_user'))
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # 获取表头
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append('')
        
        # 检查必要的列
        required_columns = ['用户名*']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            flash(f'Excel文件缺少必要的列：{", ".join(missing_columns)}', 'error')
            return redirect(url_for('create_user'))
        
        # 获取列索引
        col_indices = {}
        for i, header in enumerate(headers):
            if header in ['用户名*', '密码', '姓名', '邮箱', '角色']:
                col_indices[header] = i
        
        # 读取数据行
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(cell is not None for cell in row):  # 跳过空行
                data_rows.append(row)
        
        if len(data_rows) == 0:
            flash('Excel文件中没有有效的用户数据', 'error')
            return redirect(url_for('create_user'))
        
        if len(data_rows) > 1000:
            flash('一次最多只能导入1000个用户，请分批导入', 'error')
            return redirect(url_for('create_user'))
        
        created_users = []
        errors = []
        
        for index, row in enumerate(data_rows):
            try:
                # 获取各列数据
                username_col = col_indices.get('用户名*', 0)
                password_col = col_indices.get('密码', -1)
                name_col = col_indices.get('姓名', -1)
                email_col = col_indices.get('邮箱', -1)
                role_col = col_indices.get('角色', -1)
                
                username = str(row[username_col]).strip() if row[username_col] else ''
                # 检查原始密码是否为空（用于判断是否使用默认密码）
                original_password = str(row[password_col]).strip() if password_col >= 0 and len(row) > password_col and row[password_col] else ''
                password = original_password if original_password and original_password != 'nan' else '123456'
                name = str(row[name_col]).strip() if name_col >= 0 and len(row) > name_col and row[name_col] else None
                email = str(row[email_col]).strip() if email_col >= 0 and len(row) > email_col and row[email_col] else None
                role = str(row[role_col]).strip().lower() if role_col >= 0 and len(row) > role_col and row[role_col] else 'user'
                
                # 清理数据
                if name == '' or name == 'nan':
                    name = None
                if email == '' or email == 'nan':
                    email = None
                if role == '' or role == 'nan':
                    role = 'user'
                
                # 清理密码输入
                password = sanitize_password_input(password)
                
                # 验证用户名
                if not username:
                    errors.append(f'第{index+2}行：用户名不能为空')
                    continue
                
                # 验证用户名唯一性
                if User.query.filter_by(username=username, is_deleted=False).first():
                    errors.append(f'第{index+2}行：用户名 {username} 已存在')
                    continue
                
                # 只有当管理员提供了密码时才进行密码安全性检查
                # 如果密码字段为空，则使用默认密码并跳过密码检查
                is_using_default_password = not original_password or original_password == 'nan'
                if not is_using_default_password:
                    # 管理员提供了密码，需要进行安全性检查
                    is_valid, error_msg = validate_password_security(password)
                    if not is_valid:
                        errors.append(f'第{index+2}行：密码不符合安全要求 - {error_msg}')
                        continue
                
                # 验证角色
                if role not in ['user', 'admin']:
                    errors.append(f'第{index+2}行：角色 {role} 无效，只能是 user 或 admin')
                    continue
                
                # 验证邮箱格式（如果提供）
                if email:
                    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                    if not re.match(email_pattern, email):
                        errors.append(f'第{index+2}行：邮箱格式无效 {email}')
                        continue
                
                # 创建用户
                user = User(
                    username=username,
                    password_hash=generate_password_hash(password),
                    name=name,
                    email=email,
                    role=role,
                    created_by=current_user.id
                )
                db.session.add(user)
                created_users.append({
                    'username': username,
                    'password': password,
                    'name': name,
                    'email': email,
                    'role': role
                })
                
            except Exception as e:
                errors.append(f'第{index+2}行：处理失败 - {str(e)}')
                continue
        
        # 提交数据库事务
        if created_users:
            try:
                db.session.commit()
                flash(f'成功导入 {len(created_users)} 个用户', 'success')
                # 创建Excel导入成功消息
                user_list = '\n'.join([f"• {user['username']} ({user['name'] or '未设置姓名'})" for user in created_users])
                create_message(
                    user_id=current_user.id,
                    message_type='users_imported',
                    title=f'Excel导入成功：{len(created_users)}个用户',
                    content=f'您已成功通过Excel文件导入{len(created_users)}个用户账号\n\n导入时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n\n用户列表：\n{user_list}\n\n请及时将账号信息安全地分发给相应用户。',
                    related_type='user_management'
                )
            except Exception as e:
                db.session.rollback()
                flash(f'导入失败：{str(e)}', 'error')
                return redirect(url_for('create_user'))
        
        # 显示错误信息
        if errors:
            for error in errors[:10]:  # 最多显示10个错误
                flash(error, 'warning')
            if len(errors) > 10:
                flash(f'还有 {len(errors)-10} 个错误未显示', 'warning')
        
        if created_users:
            return render_template('user_created.html', users=created_users)
        else:
            flash('没有成功导入任何用户', 'error')
            return redirect(url_for('create_user'))
            
    except Exception as e:
        flash(f'读取Excel文件失败：{str(e)}', 'error')
        return redirect(url_for('create_user'))

@app.route('/toggle_user/<int:user_id>')
@login_required
@admin_required
def toggle_user(user_id):
    # 使用线程锁确保用户状态修改的并发安全
    with _user_operation_lock:
        try:
            with safe_db_transaction() as session:
                user = User.query.with_for_update().get_or_404(user_id)
                if user.id == current_user.id:
                    flash('不能禁用自己的账号', 'error')
                else:
                    user.is_enabled = not user.is_enabled
                    status = '启用' if user.is_enabled else '禁用'
                    flash(f'已{status}用户 {user.username}', 'success')
        except Exception as e:
            flash(f'操作失败：{str(e)}', 'error')
    
    return redirect(url_for('users'))

@app.route('/delete_user/<int:user_id>')
@login_required
@admin_required
def delete_user(user_id):
    """软删除用户 - 仅管理员可操作"""
    # 使用线程锁确保用户删除操作的并发安全
    with _user_operation_lock:
        try:
            with safe_db_transaction() as session:
                user = User.query.with_for_update().get_or_404(user_id)
                
                # 不能删除自己
                if user.id == current_user.id:
                    flash('不能删除自己的账号', 'error')
                    return redirect(url_for('users'))
                
                # 检查用户是否已被删除
                if user.is_deleted:
                    flash('用户已被删除', 'warning')
                    return redirect(url_for('users'))
                
                # 执行软删除
                user.is_deleted = True
                flash(f'已删除用户 {user.username}', 'success')
                
        except Exception as e:
            flash(f'删除失败：{str(e)}', 'error')
    
    return redirect(url_for('users'))

@app.route('/print')
@login_required
@first_login_required
def print_page():
    return render_template('print.html', template_mapping=TEMPLATE_MAPPING)

@app.route('/search_student')
@login_required
def search_student():
    student_code = request.args.get('student_code', '').strip()
    
    if not student_code:
        return jsonify({'error': '请输入学员编码'}), 400
    
    # 尝试使用实际的学员凭证搜索功能（包含班级凭证和充值提现记录）
    try:
        from utils.certificate_processors.search_student_certificate import search_student
        
        # 从数据库获取活跃的cookies配置
        cookies = None
        active_config = CookiesConfig.query.filter_by(is_active=True).first()
        if active_config:
            try:
                cookies = json.loads(active_config.cookies_data)
            except:
                pass
        
        # 调用实际的搜索功能
        search_result = search_student(cookies, current_user, student_code)
        
        if search_result == 0:
            return jsonify({'error': '未找到该学员的信息'}), 404
        elif search_result == 404:
            # API调用失败，返回错误信息并提醒管理员更新cookies
            error_msg = 'API调用失败，可能是网络问题或认证失效。'
            if current_user.role != 'admin':
                error_msg += '请联系管理员更新系统认证配置。'
            else:
                error_msg += '请前往Cookies配置页面更新认证信息。'
            
            return jsonify({
                'error': error_msg,
                'need_admin_attention': True,
                'is_admin': current_user.role == 'admin'
            }), 404
        elif isinstance(search_result, dict) and 'reports' in search_result:
            # 成功获取到实际数据（新格式）
            student_info = search_result
            
            # 为每个报告添加详细信息描述
            for report in student_info.get('reports', []):
                biz_type = report.get('biz_type')
                data = report.get('data', {})
                
                # 生成详细信息描述
                _, detail_info = _get_certificate_info(biz_type, data)
                report['description'] = detail_info
            
            return jsonify(student_info)
        else:
            return jsonify({'error': '未找到该学员的凭证信息'}), 404
            
    except Exception as e:
        print(f"搜索学生信息时发生错误: {str(e)}")
        # 发生错误时回退到模拟数据
        return _fallback_mock_search(student_code)


def _fallback_mock_search(student_code):
    """回退到模拟数据搜索"""
    # 不再提供测试数据，直接返回未找到
    return jsonify({'error': '未找到该学员的信息'}), 404

@app.route('/generate_print', methods=['POST'])
@login_required
def generate_print():
    # 使用线程锁确保打印操作的并发安全
    with _print_lock:
        try:
            data = request.json
            biz_type = data.get('biz_type')
            student_data = data.get('student_data')
            
            if not biz_type or not student_data:
                return jsonify({'error': '缺少必要参数'}), 400
            
            # 使用新的凭证管理器生成打印图像
            # 方法1: 使用新的独立处理器（推荐）
            try:
                image_path = generate_certificate_by_type(biz_type, student_data)
            except Exception as e:
                print(f"新处理器失败: {str(e)}")
                image_path = None
            
            # 方法2: 如果新处理器失败，使用原有方法（备选）
            if not image_path:
                try:
                    image_path = print_certificate_by_biz_type(biz_type, student_data)
                except Exception as e:
                    print(f"原有方法也失败: {str(e)}")
                    image_path = None
            
            # 方法3: 最后的传统方法（保底）
            if not image_path:
                # 创建打印消息
                message = {
                    "PrintType": "proofprintnew",
                    "Info": {
                        "Params": {
                            "BizType": biz_type,
                            "JsonString": json.dumps(student_data),
                            "DefaultPrinter": "",
                            "DefaultPrintNumber": 1,
                            "NeedPreview": True,
                            "SchoolId": 35,
                            "CurrencySymbol": "¥"
                        }
                    }
                }
                
                # 生成打印图像
                simulator = ProofPrintSimulator()
                image_path = simulator.process_print_request(message)
            
            if image_path and os.path.exists(image_path):
                # 将图像转换为base64
                with open(image_path, 'rb') as img_file:
                    img_data = base64.b64encode(img_file.read()).decode()
                
                # 使用并发安全的文件名生成器
                filename = generate_unique_filename("print_certificate", "png")
                
                # 根据biz_type确定凭证名称和详细信息
                biz_name, detail_info = _get_certificate_info(biz_type, student_data)
                
                # 使用安全的数据库事务上下文
                with safe_db_transaction() as session:
                    print_log = PrintLog(
                        user_id=current_user.id,
                        student_code=student_data.get('sStudentCode', ''),
                        student_name=student_data.get('sStudentName', ''),
                        biz_type=biz_type,
                        biz_name=biz_name,
                        print_data=json.dumps(student_data, ensure_ascii=False),
                        detail_info=detail_info
                    )
                    session.add(print_log)
                    session.flush()  # 获取print_log.id
                    
                    # 创建打印成功消息
                    message_content = f'凭证打印成功！\n\n学员信息：\n学员编码：{student_data.get("sStudentCode", "")}\n学员姓名：{student_data.get("sStudentName", "")}\n\n凭证信息：\n凭证类型：{biz_name}'
                    if detail_info:
                        message_content += f'\n详细信息：{detail_info}'
                    
                    message = Message(
                        user_id=current_user.id,
                        message_type='print_success',
                        title=f'{biz_name}打印成功',
                        content=message_content,
                        related_id=print_log.id,
                        related_type='print_log'
                    )
                    session.add(message)
                
                # 清理临时文件
                try:
                    os.remove(image_path)
                    json_file = image_path.replace('.png', '.json')
                    if os.path.exists(json_file):
                        os.remove(json_file)
                except:
                    pass
                
                return jsonify({
                    'success': True,
                    'image': img_data,
                    'filename': filename
                })
            else:
                return jsonify({'error': '打印处理失败'}), 500
                
        except Exception as e:
            return jsonify({'error': f'生成打印失败：{str(e)}'}), 500

@app.route('/api/certificate_types')
@login_required  
def api_certificate_types():
    """获取所有可用的凭证类型信息"""
    try:
        certificates = get_available_certificates()
        
        # 转换为前端友好的格式
        result = []
        for cert_type, config in certificates.items():
            for i, template in enumerate(config['templates']):
                biz_type = config['biz_types'][i] if i < len(config['biz_types']) else config['biz_types'][0]
                result.append({
                    'biz_type': biz_type,
                    'template_name': template,
                    'category': cert_type,
                    'processor': config['processor'],
                    'display_name': template.replace('.mrt', '')
                })
        
        return jsonify({
            'success': True,
            'certificates': result,
            'template_mapping': TEMPLATE_MAPPING
        })
    except Exception as e:
        return jsonify({'error': f'获取凭证类型失败：{str(e)}'}), 500

@app.route('/print_logs')
@login_required
def print_logs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    # 限制每页记录数量范围
    if per_page not in [10, 20, 50, 100]:
        per_page = 20
    
    # 搜索参数
    search_student = request.args.get('search_student', '').strip()  # 学员编码或姓名
    search_biz_type = request.args.get('search_biz_type', '', type=str)  # 凭证类型
    search_date_start = request.args.get('search_date_start', '').strip()  # 开始日期
    search_date_end = request.args.get('search_date_end', '').strip()  # 结束日期
    
    # 排序参数
    sort_by = request.args.get('sort_by', 'print_time')  # 默认按打印时间排序
    sort_order = request.args.get('sort_order', 'desc')  # 默认降序
    
    # 构建查询
    query = PrintLog.query.filter_by(is_deleted=False)  # 只显示未删除的记录
    
    # 权限控制
    if current_user.role != 'admin':
        # 普通用户只能查看自己的日志
        query = query.filter_by(user_id=current_user.id)
    
    # 学员信息筛选
    if search_student:
        query = query.filter(
            db.or_(
                PrintLog.student_code.like(f'%{search_student}%'),
                PrintLog.student_name.like(f'%{search_student}%')
            )
        )
    
    # 凭证类型筛选
    if search_biz_type:
        try:
            biz_type_int = int(search_biz_type)
            query = query.filter(PrintLog.biz_type == biz_type_int)
        except ValueError:
            # 如果不是数字，按凭证名称搜索
            query = query.filter(PrintLog.biz_name.like(f'%{search_biz_type}%'))
    
    # 时间范围筛选
    if search_date_start:
        try:
            start_date = datetime.strptime(search_date_start, '%Y-%m-%d')
            query = query.filter(PrintLog.print_time >= start_date)
        except ValueError:
            pass
    
    if search_date_end:
        try:
            end_date = datetime.strptime(search_date_end, '%Y-%m-%d')
            # 结束日期包含当天，所以加1天
            end_date = end_date + timedelta(days=1)
            query = query.filter(PrintLog.print_time < end_date)
        except ValueError:
            pass
    
    # 排序和分页
    # 构建排序
    valid_sort_fields = {
        'print_time': PrintLog.print_time,
        'student_code': PrintLog.student_code,
        'student_name': PrintLog.student_name,
        'biz_type': PrintLog.biz_type,
        'biz_name': PrintLog.biz_name,
        'user_id': PrintLog.user_id,
        'id': PrintLog.id
    }
    
    sort_field = valid_sort_fields.get(sort_by, PrintLog.print_time)
    
    if sort_order == 'asc':
        query = query.order_by(sort_field.asc())
    else:
        query = query.order_by(sort_field.desc())
    
    logs = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # 获取凭证类型列表用于下拉框
    biz_types = db.session.query(PrintLog.biz_type, PrintLog.biz_name).filter_by(is_deleted=False).distinct().all()
    
    # 构建搜索参数字典，用于分页链接保持搜索状态
    search_params = {}
    if search_student:
        search_params['search_student'] = search_student
    if search_biz_type:
        search_params['search_biz_type'] = search_biz_type
    if search_date_start:
        search_params['search_date_start'] = search_date_start
    if search_date_end:
        search_params['search_date_end'] = search_date_end
    
    # 添加排序和分页参数
    search_params['sort_by'] = sort_by
    search_params['sort_order'] = sort_order
    search_params['per_page'] = per_page
    
    # 构建分页URL参数字符串
    search_query_string = ''
    if search_params:
        query_parts = []
        for key, value in search_params.items():
            if value:  # 只包含非空值
                query_parts.append(f'{key}={value}')
        if query_parts:
            search_query_string = '&' + '&'.join(query_parts)
    
    return render_template('print_logs.html', 
                         logs=logs, 
                         biz_types=biz_types,
                         search_params=search_params,
                         search_query_string=search_query_string,
                         current_per_page=per_page)

@app.route('/delete_print_log/<int:log_id>')
@login_required
@admin_required
def delete_print_log(log_id):
    """删除打印记录 - 仅管理员可操作"""
    # 使用线程锁确保打印记录删除的并发安全
    with _print_lock:
        try:
            with safe_db_transaction() as session:
                log = PrintLog.query.with_for_update().get_or_404(log_id)
                log.is_deleted = True  # 软删除
                flash('打印记录已删除', 'success')
        except Exception as e:
            flash(f'删除失败：{str(e)}', 'error')
    
    return redirect(url_for('print_logs'))

@app.route('/first_login_change_password', methods=['GET', 'POST'])
@login_required
def first_login_change_password():
    # 如果不是首次登录，重定向到普通密码修改页面
    if not current_user.is_first_login:
        return redirect(url_for('change_password'))
    
    if request.method == 'POST':
        # 清理输入数据，防止注入
        current_password = sanitize_password_input(request.form.get('current_password', ''))
        new_password = sanitize_password_input(request.form.get('new_password', ''))
        confirm_password = sanitize_password_input(request.form.get('confirm_password', ''))
        captcha_input = request.form.get('captcha', '').upper()
        
        # 验证验证码
        if 'captcha_code' not in session or captcha_input != session.get('captcha_code'):
            flash('验证码错误', 'error')
            # 清除旧的验证码
            session.pop('captcha_code', None)
            return render_template('first_login_change_password.html')
        
        # 清除验证码（一次性使用）
        session.pop('captcha_code', None)
        
        # 验证当前密码
        if not current_password:
            flash('请输入当前密码', 'error')
            return render_template('first_login_change_password.html')
            
        if not check_password_hash(current_user.password_hash, current_password):
            flash('当前密码错误', 'error') 
            return render_template('first_login_change_password.html')
        
        # 验证新密码安全性
        is_valid, error_msg = validate_password_security(new_password)
        if not is_valid:
            flash(f'新密码不符合安全要求：{error_msg}', 'error')
            return render_template('first_login_change_password.html')
        
        # 验证密码确认
        if new_password != confirm_password:
            flash('两次输入的新密码不一致', 'error')
            return render_template('first_login_change_password.html')
        
        # 检查新密码是否与旧密码相同
        if check_password_hash(current_user.password_hash, new_password):
            flash('新密码不能与当前密码相同', 'error')
            return render_template('first_login_change_password.html')
        
        try:
            # 更新密码并标记非首次登录
            current_user.password_hash = generate_password_hash(new_password)
            current_user.is_first_login = False
            db.session.commit()
            
            flash('密码修改成功，欢迎使用系统！', 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'密码修改失败：{str(e)}', 'error')
            return render_template('first_login_change_password.html')
    
    return render_template('first_login_change_password.html')

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
@first_login_required
def change_password():
    if request.method == 'POST':
        # 清理输入数据，防止注入
        current_password = sanitize_password_input(request.form.get('current_password', ''))
        new_password = sanitize_password_input(request.form.get('new_password', ''))
        confirm_password = sanitize_password_input(request.form.get('confirm_password', ''))
        captcha_input = request.form.get('captcha', '').upper()
        
        # 验证验证码
        if 'captcha_code' not in session or captcha_input != session.get('captcha_code'):
            flash('验证码错误', 'error')
            # 清除旧的验证码
            session.pop('captcha_code', None)
            return render_template('change_password.html')
        
        # 清除验证码（一次性使用）
        session.pop('captcha_code', None)
        
        # 验证当前密码
        if not current_password:
            flash('请输入当前密码', 'error')
            return render_template('change_password.html')
            
        if not check_password_hash(current_user.password_hash, current_password):
            flash('当前密码错误', 'error') 
            return render_template('change_password.html')
        
        # 验证新密码安全性
        is_valid, error_msg = validate_password_security(new_password)
        if not is_valid:
            flash(f'新密码不符合安全要求：{error_msg}', 'error')
            return render_template('change_password.html')
        
        # 验证密码确认
        if new_password != confirm_password:
            flash('两次输入的新密码不一致', 'error')
            return render_template('change_password.html')
        
        # 检查新密码是否与旧密码相同
        if check_password_hash(current_user.password_hash, new_password):
            flash('新密码不能与当前密码相同', 'error')
            return render_template('change_password.html')
        
        try:
            # 更新密码
            current_user.password_hash = generate_password_hash(new_password)
            db.session.commit()
            
            flash('密码修改成功', 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'密码修改失败：{str(e)}', 'error')
            return render_template('change_password.html')
    
    return render_template('change_password.html')

@app.route('/edit_profile', methods=['GET', 'POST'])
@login_required
@first_login_required
def edit_profile():
    """个人资料编辑"""
    if request.method == 'POST':
        new_username = request.form.get('username', '').strip()
        new_name = request.form.get('name', '').strip()
        new_email = request.form.get('email', '').strip()
        reason = request.form.get('reason', '').strip()
        
        # 检查是否有变更
        changes = {}
        if new_username != current_user.username:
            # 检查用户名是否已存在
            existing_user = User.query.filter_by(username=new_username, is_deleted=False).first()
            if existing_user and existing_user.id != current_user.id:
                flash('用户名已存在，请选择其他用户名', 'error')
                return render_template('edit_profile.html')
            changes['username'] = new_username
            
        if new_name != (current_user.name or ''):
            changes['name'] = new_name
            
        if new_email != (current_user.email or ''):
            changes['email'] = new_email
        
        if not changes:
            flash('没有检测到任何变更', 'info')
            return render_template('edit_profile.html')
        
        if not reason:
            flash('请填写变更理由', 'error')
            return render_template('edit_profile.html')
        
        # 管理员可以直接修改姓名和邮箱
        if current_user.role == 'admin':
            if 'name' in changes:
                current_user.name = changes['name']
            if 'email' in changes:
                current_user.email = changes['email']
            
            # 如果管理员要修改用户名，仍需要审批
            if 'username' in changes:
                change_request = UserChangeRequest(
                    user_id=current_user.id,
                    request_type='profile',
                    new_username=changes['username'],
                    old_username=current_user.username,
                    reason=reason,
                    status='pending'
                )
                db.session.add(change_request)
                db.session.flush()  # 获取change_request.id
                
                # 创建提交申请消息
                create_message(
                    user_id=current_user.id,
                    message_type='change_request_submitted',
                    title='信息变更申请已提交',
                    content=f'您的用户名变更申请已提交，等待其他管理员审批。\n\n变更内容：\n用户名: {current_user.username} → {changes["username"]}\n\n申请理由：{reason}',
                    related_id=change_request.id,
                    related_type='change_request'
                )
                
                flash('用户名变更申请已提交，等待其他管理员审批', 'info')
            else:
                flash('个人信息修改成功', 'success')
            
            db.session.commit()
        else:
            # 普通用户需要管理员审批
            change_request = UserChangeRequest(
                user_id=current_user.id,
                request_type='profile',
                new_username=changes.get('username'),
                new_name=changes.get('name'),
                new_email=changes.get('email'),
                old_username=current_user.username,
                old_name=current_user.name,
                old_email=current_user.email,
                reason=reason,
                status='pending'
            )
            db.session.add(change_request)
            db.session.flush()  # 获取change_request.id
            
            # 创建提交申请消息
            change_items = []
            if 'username' in changes:
                change_items.append(f"用户名: {current_user.username} → {changes['username']}")
            if 'name' in changes:
                old_name = current_user.name or "未设置"
                change_items.append(f"姓名: {old_name} → {changes['name']}")
            if 'email' in changes:
                old_email = current_user.email or "未设置"
                change_items.append(f"邮箱: {old_email} → {changes['email']}")
            
            message_content = f'您的信息变更申请已提交，等待管理员审批。\n\n变更内容：\n' + '\n'.join(change_items) + f'\n\n申请理由：{reason}'
            
            create_message(
                user_id=current_user.id,
                message_type='change_request_submitted',
                title='信息变更申请已提交',
                content=message_content,
                related_id=change_request.id,
                related_type='change_request'
            )
            
            db.session.commit()
            flash('个人信息变更申请已提交，等待管理员审批', 'info')
        
        return redirect(url_for('edit_profile'))
    
    # 获取用户的待处理申请
    pending_requests = UserChangeRequest.query.filter_by(
        user_id=current_user.id, 
        status='pending'
    ).order_by(UserChangeRequest.created_at.desc()).all()
    
    return render_template('edit_profile.html', pending_requests=pending_requests)

@app.route('/messages')
@login_required
@first_login_required
def messages():
    """消息盒子页面"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    # 限制每页记录数量范围
    if per_page not in [10, 20, 50]:
        per_page = 20
    
    # 搜索和筛选参数
    message_type = request.args.get('message_type', '').strip()
    is_read = request.args.get('is_read', '').strip()
    
    # 构建查询
    query = Message.query.filter_by(user_id=current_user.id)
    
    # 消息类型筛选
    if message_type:
        query = query.filter(Message.message_type == message_type)
    
    # 已读状态筛选
    if is_read == 'true':
        query = query.filter(Message.is_read == True)
    elif is_read == 'false':
        query = query.filter(Message.is_read == False)
    
    # 按创建时间降序排列
    query = query.order_by(Message.created_at.desc())
    
    # 分页
    messages_pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # 获取统计信息
    total_messages = Message.query.filter_by(user_id=current_user.id).count()
    unread_messages = Message.query.filter_by(user_id=current_user.id, is_read=False).count()
    
    stats = {
        'total': total_messages,
        'unread': unread_messages,
        'read': total_messages - unread_messages
    }
    
    return render_template('messages.html', 
                         messages=messages_pagination.items,
                         pagination=messages_pagination,
                         stats=stats,
                         message_type=message_type,
                         is_read=is_read,
                         per_page=per_page)

@app.route('/api/mark_message_read/<int:message_id>', methods=['POST'])
@login_required
def mark_message_read(message_id):
    """标记消息为已读"""
    # 使用线程锁确保消息操作的并发安全
    with _message_operation_lock:
        try:
            with safe_db_transaction() as session:
                message = Message.query.with_for_update().get_or_404(message_id)
                
                # 检查权限
                if message.user_id != current_user.id:
                    return jsonify({'error': '无权限操作此消息'}), 403
                
                if not message.is_read:
                    message.is_read = True
                    message.read_at = get_beijing_datetime()
                
                return jsonify({'success': True})
        except Exception as e:
            return jsonify({'error': f'操作失败：{str(e)}'}), 500

@app.route('/api/mark_all_messages_read', methods=['POST'])
@login_required
def mark_all_messages_read():
    """标记所有消息为已读"""
    # 使用线程锁确保批量消息操作的并发安全
    with _message_operation_lock:
        try:
            with safe_db_transaction() as session:
                # 使用SELECT FOR UPDATE确保一致性
                unread_messages = Message.query.filter_by(
                    user_id=current_user.id, 
                    is_read=False
                ).with_for_update().all()
                
                for message in unread_messages:
                    message.is_read = True
                    message.read_at = get_beijing_datetime()
                
                return jsonify({'success': True, 'updated_count': len(unread_messages)})
        except Exception as e:
            return jsonify({'error': f'操作失败：{str(e)}'}), 500

@app.route('/api/delete_message/<int:message_id>', methods=['DELETE'])
@login_required
def delete_message(message_id):
    """删除消息"""
    # 使用线程锁确保消息删除的并发安全
    with _message_operation_lock:
        try:
            with safe_db_transaction() as session:
                message = Message.query.with_for_update().get_or_404(message_id)
                
                # 检查权限
                if message.user_id != current_user.id:
                    return jsonify({'error': '无权限操作此消息'}), 403
                
                session.delete(message)
                return jsonify({'success': True})
        except Exception as e:
            return jsonify({'error': f'删除失败：{str(e)}'}), 500

@app.route('/api/get_unread_count')
@login_required
def get_unread_count():
    unread_count = Message.query.filter_by(user_id=current_user.id, is_read=False).count()
    return jsonify({'unread_count': unread_count})

@app.route('/api/check_password_strength', methods=['POST'])
@login_required
def api_check_password_strength():
    """检查密码强度的API接口"""
    try:
        data = request.get_json()
        password = data.get('password', '')
        
        # 清理密码输入
        password = sanitize_password_input(password)
        
        # 验证密码安全性
        is_valid, error_msg = validate_password_security(password)
        
        # 检查密码强度
        strength_level, strength_text, suggestions = check_password_strength(password)
        
        return jsonify({
            'is_valid': is_valid,
            'error_message': error_msg if not is_valid else None,
            'strength_level': strength_level,
            'strength_text': strength_text,
            'suggestions': suggestions,
            'score': strength_level * 20  # 转换为百分比
        })
        
    except Exception as e:
        return jsonify({
            'is_valid': False,
            'error_message': f'检查失败：{str(e)}',
            'strength_level': 0,
            'strength_text': '无',
            'suggestions': ['请输入有效密码'],
            'score': 0
        }), 500

@app.route('/api/version')
def get_version():
    """获取系统版本信息"""
    return jsonify({
        'name': '南昌新东方凭证打印系统',
        'version': '2.4.0',
        'release_date': '2025-06-20',
        'description': '支持多种凭证打印、用户管理、Excel批量导入、消息通知的综合管理系统'
    })

@app.route('/health')
def health_check():
    """健康检查端点 - 用于Kubernetes存活检查"""
    return jsonify({
        'status': 'healthy',
                    'timestamp': get_beijing_time_str(),
        'service': '南昌新东方凭证打印系统'
    })

@app.route('/healthz')
def health_check_k8s():
    """Kubernetes标准健康检查端点"""
    try:
        # 简单的数据库连接检查
        db.session.execute('SELECT 1')
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': get_beijing_time_str()
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'database': 'disconnected',
            'error': str(e),
            'timestamp': get_beijing_time_str()
        }), 503

@app.route('/ready')
def readiness_check():
    """就绪检查端点 - 检查应用是否准备好接收流量"""
    try:
        # 检查数据库连接
        db.session.execute('SELECT 1')
        
        # 检查基本功能
        return jsonify({
            'status': 'ready',
            'database': 'connected',
            'timestamp': get_beijing_time_str()
        })
    except Exception as e:
        return jsonify({
            'status': 'not_ready',
            'database': 'disconnected',
            'error': str(e),
            'timestamp': get_beijing_time_str()
        }), 503

@app.route('/tech_support')
@login_required
def tech_support():
    """技术支持页面"""
    return render_template('tech_support.html')

@app.route('/debug_cookies')
@login_required
def debug_cookies():
    """调试：显示当前所有cookies"""
    cookies_info = {}
    
    for cookie_name, cookie_value in request.cookies.items():
        cookies_info[cookie_name] = cookie_value
        
        # 特别处理login_time cookie
        if cookie_name == 'login_time':
            try:
                timestamp = float(cookie_value)
                login_time = datetime.fromtimestamp(timestamp)
                now = datetime.now()
                elapsed = now - login_time
                remaining = timedelta(hours=12) - elapsed
                
                cookies_info[f'{cookie_name}_详细信息'] = {
                    '原始时间戳': timestamp,
                    '登录时间': login_time.strftime('%Y-%m-%d %H:%M:%S'),
                    '当前时间': now.strftime('%Y-%m-%d %H:%M:%S'),
                    '已使用时间': f'{elapsed.total_seconds()/3600:.2f} 小时',
                    '剩余时间': f'{remaining.total_seconds()/3600:.2f} 小时',
                    '是否即将过期': remaining.total_seconds() < 3600,
                    '过期时间': (login_time + timedelta(hours=12)).strftime('%Y-%m-%d %H:%M:%S')
                }
            except:
                cookies_info[f'{cookie_name}_详细信息'] = '解析失败'
    
    # 用户身份验证状态
    login_user_cookie = request.cookies.get('login_user')
    login_user_id_cookie = request.cookies.get('login_user_id')
    
    identity_check = {
        'cookie中的用户名': login_user_cookie,
        'cookie中的用户ID': login_user_id_cookie,
        '当前登录用户名': current_user.username,
        '当前登录用户ID': current_user.id,
        '用户名匹配': login_user_cookie == current_user.username if login_user_cookie else False,
        '用户ID匹配': str(login_user_id_cookie) == str(current_user.id) if login_user_id_cookie else False,
        '身份验证通过': (login_user_cookie == current_user.username and 
                      str(login_user_id_cookie) == str(current_user.id)) if login_user_cookie and login_user_id_cookie else False
    }
    
    return jsonify({
        '当前cookies': cookies_info,
        '用户身份验证': identity_check,
        '总cookies数量': len(request.cookies),
        '检查时间': get_beijing_time_str()
    })



@app.route('/cookies_config')
@login_required
@admin_required
def cookies_config():
    """Cookies配置管理页面"""
    configs = CookiesConfig.query.order_by(CookiesConfig.updated_at.desc()).all()
    active_config = CookiesConfig.query.filter_by(is_active=True).first()
    return render_template('cookies_config.html', configs=configs, active_config=active_config)

# 在save_cookies路由之前添加cookies数据解析函数
def parse_cookies_data(cookies_input):
    """
    解析不同格式的cookies数据
    支持以下格式：
    1. JSON格式
    2. Python字典格式（用单引号）
    3. 键值对格式（换行分隔）
    4. 键值对格式（分号分隔）
    5. 浏览器原始Cookie字符串格式
    """
    cookies_input = cookies_input.strip()
    
    if not cookies_input:
        return {}
    
    # 尝试解析JSON格式
    try:
        if cookies_input.startswith('{') and cookies_input.endswith('}'):
            return json.loads(cookies_input)
    except json.JSONDecodeError:
        pass
    
    # 尝试解析Python字典格式（将单引号替换为双引号）
    try:
        if cookies_input.startswith('{') and cookies_input.endswith('}'):
            # 处理Python字典格式：将单引号替换为双引号
            # 但要注意值中可能包含单引号，所以使用正则表达式处理
            import re
            # 替换键名的单引号为双引号
            cookies_input_fixed = re.sub(r"'([^']*?)':", r'"\1":', cookies_input)
            # 替换值的单引号为双引号（但保留转义的单引号）
            cookies_input_fixed = re.sub(r": '([^']*?)'([,}])", r': "\1"\2', cookies_input_fixed)
            return json.loads(cookies_input_fixed)
    except (json.JSONDecodeError, Exception):
        pass
    
    # 处理重复键的特殊情况：手动解析字典格式
    try:
        if cookies_input.startswith('{') and cookies_input.endswith('}'):
            cookies_dict = {}
            content = cookies_input[1:-1].strip()  # 去掉大括号
            
            # 使用正则表达式匹配键值对
            import re
            # 匹配 'key': 'value' 或 "key": "value" 格式
            pattern = r"""['"]([^'"]*?)['"]:\s*['"]([^'"]*?)['"]"""
            matches = re.findall(pattern, content)
            
            for key, value in matches:
                # 处理重复键：如果键已存在，在键名后添加数字后缀
                original_key = key
                counter = 1
                while key in cookies_dict:
                    key = f"{original_key}_{counter}"
                    counter += 1
                cookies_dict[key] = value
            
            if cookies_dict:
                return cookies_dict
    except Exception as e:
        print(f"Debug: Manual parsing failed: {e}")
        pass
    
    # 尝试使用eval解析Python字典（谨慎使用）
    try:
        if cookies_input.startswith('{') and cookies_input.endswith('}'):
            # 安全性检查：只允许基本的字典语法
            if not any(dangerous in cookies_input.lower() for dangerous in ['import', 'exec', 'eval', '__']):
                result = eval(cookies_input)
                if isinstance(result, dict):
                    return result
    except Exception:
        pass
    
    # 解析键值对格式
    cookies_dict = {}
    
    # 判断分隔符
    if cookies_input.startswith('{') and cookies_input.endswith('}'):
        # 去掉大括号，然后按逗号分割
        content = cookies_input[1:-1].strip()
        if ',' in content:
            pairs = content.split(',')
        else:
            pairs = [content]
    elif ';' in cookies_input and '\n' not in cookies_input:
        # 分号分隔格式
        pairs = cookies_input.split(';')
    else:
        # 换行分隔格式
        pairs = cookies_input.split('\n')
    
    for pair in pairs:
        pair = pair.strip()
        if ':' in pair and '=' not in pair:
            # 处理 'key': 'value' 格式
            key, value = pair.split(':', 1)
            key = key.strip().strip("'\"")
            value = value.strip().strip("'\"").rstrip(',')
            if key and value:
                cookies_dict[key] = value
        elif '=' in pair:
            # 处理 key=value 格式
            key, value = pair.split('=', 1)
            key = key.strip().strip("'\"")
            value = value.strip().strip("'\"").rstrip(',')
            if key and value:
                cookies_dict[key] = value
    
    return cookies_dict

@app.route('/save_cookies', methods=['POST'])
@login_required
@admin_required
def save_cookies():
    # 使用线程锁确保Cookies配置操作的并发安全
    with _cookies_config_lock:
        try:
            # 判断请求类型并获取数据
            if request.is_json:
                # JSON请求
                data = request.json
                name = data.get('name', 'ERP Cookies')
                cookies_input = data.get('cookies_data', {})
                if isinstance(cookies_input, dict):
                    cookies_data = cookies_input
                else:
                    cookies_data = parse_cookies_data(str(cookies_input))
            else:
                # 表单请求
                name = request.form.get('config_name', 'ERP Cookies')
                cookies_input = request.form.get('cookies_data', '')
                cookies_data = parse_cookies_data(cookies_input)
                
                # 添加调试信息
                print(f"Debug: cookies_input length: {len(cookies_input)}")
                print(f"Debug: cookies_data: {cookies_data}")
                print(f"Debug: cookies_data type: {type(cookies_data)}")
                print(f"Debug: cookies_data empty: {not cookies_data}")
            
            if not cookies_data:
                error_msg = f'Cookies数据解析失败或为空。输入数据长度: {len(cookies_input)}'
                if request.is_json:
                    return jsonify({'error': error_msg}), 400
                else:
                    flash(error_msg, 'error')
                    return redirect(url_for('cookies_config'))
            
            # 使用安全的数据库事务上下文
            with safe_db_transaction() as session:
                # 创建新的Cookies配置
                config = CookiesConfig(
                    name=name,
                    cookies_data=json.dumps(cookies_data, ensure_ascii=False),
                    created_by=current_user.id,
                    is_active=False  # 新创建的配置默认不激活
                )
                session.add(config)
                session.flush()  # 获取配置ID
                
                # 创建消息通知
                create_message(
                    user_id=current_user.id,
                    message_type='cookies_saved',
                    title='Cookies配置保存成功',
                    content=f'您已成功保存新的Cookies配置：{name}\n\n配置ID：{config.id}\n保存时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n\n配置已保存但尚未激活，请在测试成功后手动激活。',
                    related_id=config.id,
                    related_type='cookies_config'
                )
                
                if request.is_json:
                    return jsonify({
                        'success': True,
                        'message': 'Cookies配置保存成功',
                        'config_id': config.id
                    })
                else:
                    flash(f'Cookies配置 "{name}" 保存成功', 'success')
                    return redirect(url_for('cookies_config'))
                
        except json.JSONDecodeError:
            error_msg = 'Cookies数据格式不正确'
            if request.is_json:
                return jsonify({'error': error_msg}), 400
            else:
                flash(error_msg, 'error')
                return redirect(url_for('cookies_config'))
        except Exception as e:
            error_msg = f'保存失败：{str(e)}'
            if request.is_json:
                return jsonify({'error': error_msg}), 500
            else:
                flash(error_msg, 'error')
                return redirect(url_for('cookies_config'))

@app.route('/test_cookies/<int:config_id>')
@login_required
@admin_required
def test_cookies(config_id):
    """测试cookies配置"""
    # 使用线程锁确保配置测试的并发安全
    with _cookies_operation_lock:
        try:
            with safe_db_transaction() as session:
                config = CookiesConfig.query.with_for_update().get_or_404(config_id)
                cookies_dict = json.loads(config.cookies_data)
                
                # 使用测试学员编号测试API
                from utils.certificate_processors.search_student_certificate import search_student
                test_result = search_student(cookies_dict, current_user, 'NC24048S6UzC')  # 使用一个测试学员编号
                
                # 更新测试状态
                config.last_test_time = get_beijing_datetime()
                
                if test_result == 404:
                    config.test_status = '失败'
                    flash(f'Cookies测试失败：API请求失败', 'error')
                    # 创建测试失败消息
                    create_message(
                        user_id=current_user.id,
                        message_type='cookies_test_failed',
                        title='Cookies配置测试失败',
                        content=f'Cookies配置"{config.name}"测试失败\n\n配置ID：{config.id}\n测试时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n失败原因：API请求失败，可能是网络问题或认证失效\n\n请检查配置是否正确或网络连接是否正常。',
                        related_id=config.id,
                        related_type='cookies_config'
                    )
                elif test_result == 0:
                    config.test_status = '成功'
                    flash(f'Cookies测试成功：API连接正常（测试学员不存在是正常的）', 'success')
                    # 创建测试成功消息
                    create_message(
                        user_id=current_user.id,
                        message_type='cookies_test_success',
                        title='Cookies配置测试成功',
                        content=f'Cookies配置"{config.name}"测试成功\n\n配置ID：{config.id}\n测试时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n测试结果：API连接正常（测试学员不存在是正常的）\n\n配置可以正常使用，您可以考虑激活此配置。',
                        related_id=config.id,
                        related_type='cookies_config'
                    )
                elif isinstance(test_result, dict):
                    config.test_status = '成功'
                    flash(f'Cookies测试成功：找到学员数据', 'success')
                    # 创建测试成功消息
                    create_message(
                        user_id=current_user.id,
                        message_type='cookies_test_success',
                        title='Cookies配置测试成功',
                        content=f'Cookies配置"{config.name}"测试成功\n\n配置ID：{config.id}\n测试时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n测试结果：找到学员数据，API工作正常\n\n配置可以正常使用，您可以考虑激活此配置。',
                        related_id=config.id,
                        related_type='cookies_config'
                    )
                else:
                    config.test_status = '失败'
                    flash(f'Cookies测试失败：返回异常结果', 'error')
                    # 创建测试失败消息
                    create_message(
                        user_id=current_user.id,
                        message_type='cookies_test_failed',
                        title='Cookies配置测试失败',
                        content=f'Cookies配置"{config.name}"测试失败\n\n配置ID：{config.id}\n测试时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n失败原因：返回异常结果\n返回内容：{str(test_result)}\n\n请检查配置是否正确。',
                        related_id=config.id,
                        related_type='cookies_config'
                    )
                
        except Exception as e:
            flash(f'测试失败：{str(e)}', 'error')
    
    return redirect(url_for('cookies_config'))

@app.route('/activate_cookies/<int:config_id>')
@login_required
@admin_required
def activate_cookies(config_id):
    """激活指定的cookies配置"""
    # 使用线程锁确保配置激活的并发安全
    with _cookies_operation_lock:
        try:
            with safe_db_transaction() as session:
                # 将所有配置设为非活跃
                all_configs = CookiesConfig.query.with_for_update().all()
                for config in all_configs:
                    config.is_active = False
                
                # 激活指定配置
                target_config = CookiesConfig.query.with_for_update().get_or_404(config_id)
                target_config.is_active = True
                
                # 创建激活成功消息
                create_message(
                    user_id=current_user.id,
                    message_type='cookies_activated',
                    title='Cookies配置已激活',
                    content=f'您已成功激活Cookies配置：{target_config.name}\n\n配置ID：{target_config.id}\n激活时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n\n该配置现在是系统的活跃配置，将用于所有API请求。',
                    related_id=target_config.id,
                    related_type='cookies_config'
                )
                
                flash(f'已激活配置：{target_config.name}', 'success')
                
        except Exception as e:
            flash(f'激活失败：{str(e)}', 'error')
    
    return redirect(url_for('cookies_config'))

@app.route('/delete_cookies/<int:config_id>')
@login_required
@admin_required
def delete_cookies(config_id):
    """删除cookies配置"""
    # 使用线程锁确保配置删除的并发安全
    with _cookies_operation_lock:
        try:
            with safe_db_transaction() as session:
                config = CookiesConfig.query.with_for_update().get_or_404(config_id)
                config_name = config.name
                session.delete(config)
                
                flash(f'已删除配置：{config_name}', 'success')
                
        except Exception as e:
            flash(f'删除失败：{str(e)}', 'error')
    
    return redirect(url_for('cookies_config'))

@app.route('/cookies_auto_check_config')
@login_required
@admin_required
def cookies_auto_check_config():
    """Cookies自动检测配置页面"""
    auto_check_config = CookiesAutoCheck.query.first()
    if not auto_check_config:
        # 创建默认配置
        auto_check_config = CookiesAutoCheck(
            is_enabled=False,
            check_interval=30
        )
        db.session.add(auto_check_config)
        db.session.commit()
    
    return jsonify({
        'is_enabled': auto_check_config.is_enabled,
        'check_interval': auto_check_config.check_interval,
        'last_check_time': auto_check_config.last_check_time.strftime('%Y-%m-%d %H:%M:%S') if auto_check_config.last_check_time else None,
        'consecutive_failures': auto_check_config.consecutive_failures,
        'failure_notification_sent': auto_check_config.failure_notification_sent
    })

@app.route('/update_cookies_auto_check', methods=['POST'])
@login_required
@admin_required
def update_cookies_auto_check():
    """更新Cookies自动检测配置"""
    try:
        is_enabled = request.form.get('is_enabled') == 'true'
        check_interval = int(request.form.get('check_interval', 30))
        
        # 验证检测间隔
        if check_interval < 5:
            return jsonify({'success': False, 'error': '检测间隔不能少于5分钟'})
        if check_interval > 1440:  # 24小时
            return jsonify({'success': False, 'error': '检测间隔不能超过24小时'})
        
        auto_check_config = CookiesAutoCheck.query.first()
        if not auto_check_config:
            auto_check_config = CookiesAutoCheck()
            db.session.add(auto_check_config)
        
        auto_check_config.is_enabled = is_enabled
        auto_check_config.check_interval = check_interval
        auto_check_config.updated_at = get_beijing_datetime()
        
        db.session.commit()
        
        # 根据配置启动或停止定时任务
        if is_enabled:
            start_cookies_auto_check()
            flash(f'已启用Cookies自动检测，检测间隔：{check_interval}分钟', 'success')
            # 创建启用自动检测消息
            create_message(
                user_id=current_user.id,
                message_type='cookies_auto_check_enabled',
                title='Cookies自动检测已启用',
                content=f'您已成功启用Cookies自动检测功能\n\n检测间隔：{check_interval}分钟\n配置时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n\n系统将定期检测Cookies配置的有效性，如果检测到失效将自动通知管理员。',
                related_type='cookies_auto_check'
            )
        else:
            stop_cookies_auto_check()
            flash('已禁用Cookies自动检测', 'info')
            # 创建禁用自动检测消息
            create_message(
                user_id=current_user.id,
                message_type='cookies_auto_check_disabled',
                title='Cookies自动检测已禁用',
                content=f'您已禁用Cookies自动检测功能\n\n禁用时间：{get_beijing_datetime().strftime("%Y-%m-%d %H:%M:%S")}\n\n系统将不再自动检测Cookies配置的有效性，请手动关注配置状态。',
                related_type='cookies_auto_check'
            )
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'更新配置失败：{str(e)}'})

@app.route('/manual_check_cookies')
@login_required
@admin_required
def manual_check_cookies():
    """手动触发Cookies检测"""
    try:
        # 手动执行检测
        auto_check_cookies()
        return jsonify({'success': True, 'message': '手动检测已完成'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'手动检测失败：{str(e)}'})

@app.route('/test_class_search')
@login_required
def test_class_search():
    """测试学员凭证搜索功能（包含班级凭证和充值提现记录）"""
    student_code = request.args.get('student_code', '').strip()
    
    if not student_code:
        return jsonify({'error': '请提供学生编号', 'example': '/test_class_search?student_code=NC12345678'}), 400
    
    try:
        from utils.certificate_processors.search_student_certificate import search_student
        
        # 使用None来让函数使用默认配置
        cookies = None
        
        # 调用搜索功能
        result = search_student(cookies, current_user, student_code)
        
        if result == 0:
            return jsonify({'error': '未找到该学员', 'student_code': student_code})
        elif result == 404:
            return jsonify({'error': 'API调用失败，可能是网络问题或认证失效', 'student_code': student_code})
        elif isinstance(result, dict) and 'reports' in result:
            return jsonify({
                'success': True,
                'student_code': student_code,
                'student_name': result.get('student_name', ''),
                'gender': result.get('gender', ''),
                'certificate_count': len(result['reports']),
                'certificates': result['reports']
            })
        else:
            return jsonify({'error': '未知的返回结果', 'result': result})
            
    except Exception as e:
        return jsonify({'error': f'测试失败: {str(e)}', 'student_code': student_code})

@app.route('/search_order')
@login_required
def search_order():
    """搜索订单信息（用于报班凭证）"""
    order_code = request.args.get('order_code')
    
    if not order_code:
        return jsonify({'error': '请提供订单号', 'example': '/search_order?order_code=ORD20240101001'}), 400
    
    try:
        # 这里应该连接到实际的业务系统API来搜索订单
        # 目前先返回模拟数据
        from utils.certificate_processors.enrollment_registration_certificate import create_mock_data
        
        # 创建模拟数据，但使用搜索的订单号
        mock_data = create_mock_data()
        mock_data['sOrderCode'] = order_code
        
        # 检查是否有匹配的订单（这里是模拟逻辑）
        if order_code.startswith('ORD') or order_code.startswith('TEST'):
            return jsonify({
                'success': True,
                'order_code': order_code,
                'order_data': mock_data
            })
        else:
            return jsonify({'error': '未找到该订单', 'order_code': order_code})
            
    except Exception as e:
        return jsonify({'error': f'搜索失败: {str(e)}', 'order_code': order_code})

@app.route('/generate_enrollment_registration_certificate', methods=['POST'])
@login_required
def generate_enrollment_registration_certificate():
    """生成报班凭证"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': '请提供订单数据'}), 400
        
        # 验证必要字段
        if 'sOrderCode' not in data:
            return jsonify({'error': '缺少订单号'}), 400
        
        # 使用线程锁确保打印操作的并发安全
        with _print_lock:
            from utils.certificate_processors.enrollment_registration_certificate import generate_enrollment_registration_certificate
            
            # 生成凭证
            output_path = generate_enrollment_registration_certificate(data)
            
            # 记录打印日志
            order_code = data.get('sOrderCode', 'unknown')
            student_name = data.get('Student', {}).get('sStudentName', 'unknown')
            
            # 生成详细信息
            class_count = len(data.get('ClassAndCardArray', []))
            detail_info = f"报班凭证，订单号：{order_code}，包含{class_count}个班级"
            
            print_log = PrintLog(
                user_id=current_user.id,
                student_code=order_code,  # 使用订单号作为标识
                student_name=student_name,
                biz_type=1,  # 报班凭证的BizType
                biz_name='报班凭证',
                print_data=json.dumps(data, ensure_ascii=False),
                detail_info=detail_info
            )
            
            db.session.add(print_log)
            db.session.commit()
            
            # 创建成功消息
            create_message(
                user_id=current_user.id,
                message_type='print_success',
                title='报班凭证打印成功',
                content=f'您的报班凭证已成功生成。订单号：{order_code}，学员：{student_name}',
                related_id=print_log.id,
                related_type='print_log'
            )
            
            return jsonify({
                'success': True,
                'message': '报班凭证生成成功',
                'output_path': output_path,
                'log_id': print_log.id
            })
            
    except Exception as e:
        print(f"生成报班凭证失败: {str(e)}")
        return jsonify({'error': f'生成失败: {str(e)}'}), 500

@app.route('/test_enrollment_registration_certificate')
@login_required
def test_enrollment_registration_certificate():
    """测试报班凭证生成功能"""
    try:
        from utils.certificate_processors.enrollment_registration_certificate import test_enrollment_registration_certificate
        
        # 运行测试
        output_path = test_enrollment_registration_certificate()
        
        return jsonify({
            'success': True,
            'message': '报班凭证测试成功',
            'output_path': output_path
        })
        
    except Exception as e:
        return jsonify({'error': f'测试失败: {str(e)}'}), 500



def _get_certificate_info(biz_type, student_data):
    """根据业务类型和学员数据生成凭证名称和详细信息"""
    from utils.certificate_manager import CERTIFICATE_MAPPING
    
    # 首先尝试从student_data中获取实际的凭证名称
    biz_name = student_data.get('Title', '')
    if not biz_name:
        biz_name = student_data.get('sProofName', '')
    
    # 如果还是没有获取到，使用映射表
    if not biz_name and biz_type in CERTIFICATE_MAPPING:
        biz_name = CERTIFICATE_MAPPING[biz_type]['name']
    elif not biz_name:
        biz_name = "未知凭证类型"
    
    # 生成详细信息
    detail_info = ""
    
    if biz_type == 6:  # 学员账户充值提现凭证
        # 判断是充值还是提现
        operation_type = "充值"
        pay_info = student_data.get('sPay', '')
        title = student_data.get('Title', '')
        biz_type_info = student_data.get('sBizType', '')
        
        # 更准确的判断逻辑
        if '提现' in pay_info or '提现' in title or '提现' in biz_type_info:
            operation_type = "提现"
        elif '充值' in pay_info or '充值' in title or '充值' in biz_type_info:
            operation_type = "充值"
        
        # 获取金额信息 - 改进逻辑
        amount = ""
        
        # 首先尝试从sPay字段提取金额
        if pay_info:
            import re
            # 尝试从字符串中提取金额数字
            amount_match = re.search(r'¥?(\d+(?:,\d{3})*(?:\.\d{2})?)', pay_info)
            if amount_match:
                amount_str = amount_match.group(1).replace(',', '')
                try:
                    amount_value = float(amount_str)
                    amount = f"¥{amount_value:,.2f}"
                except:
                    amount = pay_info  # 如果解析失败，使用原始字符串
            else:
                amount = pay_info  # 如果没有找到数字，使用原始字符串
        
        # 如果sPay没有金额信息，尝试dSumBalance
        if not amount and student_data.get('dSumBalance'):
            balance_info = student_data.get('dSumBalance', '')
            if isinstance(balance_info, (int, float)):
                amount = f"¥{balance_info:,.2f}"
            else:
                # 尝试从字符串中提取数字
                import re
                balance_match = re.search(r'¥?(\d+(?:,\d{3})*(?:\.\d{2})?)', str(balance_info))
                if balance_match:
                    balance_str = balance_match.group(1).replace(',', '')
                    try:
                        balance_value = float(balance_str)
                        amount = f"¥{balance_value:,.2f}"
                    except:
                        amount = str(balance_info)
        
        detail_info = f"类型：{operation_type}"
        if amount:
            detail_info += f"，金额：{amount}"
            
    elif biz_type == 1:  # 报班凭证
        class_name = student_data.get('sClassName', '')
        fee = student_data.get('dRealFee', student_data.get('dFee', ''))
        
        detail_info = f"班级：{class_name}"
        if fee:
            try:
                fee_value = float(fee)
                detail_info += f"，费用：¥{fee_value:,.2f}"
            except:
                detail_info += f"，费用：{fee}"
                
    elif biz_type == 8:  # 退费凭证
        refund_amount = student_data.get('dRefundFee', student_data.get('dFee', ''))
        
        detail_info = f"退费"
        if refund_amount:
            try:
                amount_value = float(refund_amount)
                detail_info += f"：¥{amount_value:,.2f}"
            except:
                detail_info += f"：{refund_amount}"
            
    elif biz_type == 3:  # 退班凭证
        class_name = student_data.get('sClassName', '')
        detail_info = f"退班"
        if class_name:
            detail_info += f"：{class_name}"
            
    elif biz_type == 5:  # 处理可能的biz_type=5（与biz_type=1相同逻辑）
        class_name = student_data.get('sClassName', '')
        fee = student_data.get('dRealFee', student_data.get('dFee', ''))
        
        detail_info = f"班级：{class_name}"
        if fee:
            try:
                fee_value = float(fee)
                detail_info += f"，费用：¥{fee_value:,.2f}"
            except:
                detail_info += f"，费用：{fee}"
    
    # 如果没有生成详细信息，使用默认信息
    if not detail_info:
        detail_info = f"BizType: {biz_type}"
    
    return biz_name, detail_info

def create_admin_user():
    """创建默认管理员账户"""
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(
            username='admin',
            password_hash=generate_password_hash('admin123'),
            role='admin',
            is_first_login=False  # 管理员账户不需要强制修改密码
        )
        db.session.add(admin)
        db.session.commit()
        print("默认管理员账户已创建 - 用户名: admin, 密码: admin123")

@contextmanager
def safe_db_transaction():
    """安全的数据库事务上下文管理器"""
    try:
        yield db.session
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"数据库事务回滚: {str(e)}")
        raise
    finally:
        db.session.close()

def generate_unique_filename(prefix="print", extension="png"):
    """生成唯一的文件名，避免并发冲突"""
    timestamp = get_beijing_timestamp()
    unique_id = str(uuid.uuid4())[:8]
    return f"{prefix}_{timestamp}_{unique_id}.{extension}"

def check_username_uniqueness_safe(username, exclude_user_id=None):
    """并发安全的用户名唯一性检查"""
    query = User.query.filter_by(username=username, is_deleted=False)
    if exclude_user_id:
        query = query.filter(User.id != exclude_user_id)
    
    # 使用SELECT FOR UPDATE确保读取时的一致性
    existing_user = query.with_for_update().first()
    return existing_user is None

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        create_admin_user()
        # 初始化cookies自动检测
        init_cookies_auto_check()
    
    app.run(debug=True, host='0.0.0.0', port=8080) 